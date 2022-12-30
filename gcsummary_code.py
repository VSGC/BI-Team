# -*- coding: utf-8 -*-
"""
Created on Thu Nov 10 10:52:38 2022

@author: VAIBHAV.SRIVASTAV01
"""


import pandas as pd
import os
import datetime as date
import calendar
import numpy as np
import pyodbc
from functools import reduce
calendar.month
todays_date= date.date.today()

con_dm = pyodbc.connect('DSN=GHF_BI_CONN_DM;UID=GHF_BI_CONN;PWD=Godrej@123')
con = pyodbc.connect('DSN=GHF_BI_CONN;UID=GHF_BI_CONN;PWD=Godrej@123')
con_dm_nbfc = pyodbc.connect('DSN=GHF_BI_CONN_DM_NBFC;UID=GHF_BI_CONN;PWD=Godrej@123')


ghf_loan_view=pd.read_sql("SELECT * from prod_da_db.serve.loan_view WHERE DH_RECORD_ACTIVE_FLAG = 'Y'", con_dm)
ghf_loan_view['NBFC_FLAG'] = 'N'
ghf_base_view = pd.read_sql("SELECT * from prod_da_db.serve.BASE_VIEW WHERE DH_RECORD_ACTIVE_FLAG = 'Y'", con_dm)
ghf_base_view['NBFC_FLAG'] = 'N'
ghf_disb=pd.read_sql("select * from  prod_da_db.serve.f_finance_disbursement_details WHERE DH_RECORD_ACTIVE_FLAG = 'Y';", con_dm)
ghf_disb['NBFC_FLAG'] = 'N'
ghf_ins=pd.read_sql("select * from prod_da_db.serve.insurance_view where dh_record_active_flag='Y';",con_dm)
ghf_ins=ghf_ins.groupby('LAN_ID').sum('NET_PREMIUM')
ghf_ins['INS_FLAG']='Y'
ghf_ins=ghf_ins[['INS_FLAG']]
ghf_cust=pd.read_sql("select customer_cif,sub_category,CUSTOMER_RESIDENTIAL_STATUS from prod_da_db.serve.customer_base where dh_record_active_flag='Y'; ",con_dm)
lancif=pd.read_sql("select * from prod_da_db.serve.x_ref_lan_to_cif where dh_record_active_flag='Y' and applicant_type='APPLICANT'; ",con_dm)
lancif=lancif.merge(ghf_cust,on='CUSTOMER_CIF',how='left')
lancif=lancif[['LAN_ID','SUB_CATEGORY','CUSTOMER_RESIDENTIAL_STATUS']]

gfl_loan_view=pd.read_sql("SELECT * from prod_gfl_da_db.serve.loan_view WHERE DH_RECORD_ACTIVE_FLAG = 'Y'", con_dm_nbfc)
gfl_loan_view['NBFC_FLAG'] = 'Y'
gfl_base_view = pd.read_sql("SELECT * from prod_gfl_da_db.serve.BASE_VIEW WHERE DH_RECORD_ACTIVE_FLAG = 'Y'", con_dm_nbfc)
gfl_base_view['NBFC_FLAG'] = 'Y'
gfl_disb=pd.read_sql("select * from  prod_gfl_da_db.serve.f_finance_disbursement_details WHERE DH_RECORD_ACTIVE_FLAG = 'Y';", con_dm_nbfc)
gfl_disb['NBFC_FLAG'] = 'Y'
gfl_ins=pd.read_sql("select * from prod_gfl_da_db.serve.insurance_view where dh_record_active_flag='Y';",con_dm_nbfc)
gfl_ins=gfl_ins.groupby('LAN_ID').sum('NET_PREMIUM')
gfl_ins['INS_FLAG']='Y'
gfl_ins=gfl_ins[['INS_FLAG']]
gfl_cust=pd.read_sql("select customer_cif,sub_category,CUSTOMER_RESIDENTIAL_STATUS from prod_gfl_da_db.serve.customer_base where dh_record_active_flag='Y'; ",con_dm_nbfc)
gfl_lancif=pd.read_sql("select * from prod_gfl_da_db.serve.x_ref_lan_to_cif where dh_record_active_flag='Y' and applicant_type='APPLICANT'; ",con_dm_nbfc)
gfl_lancif=gfl_lancif.merge(gfl_cust,on='CUSTOMER_CIF',how='left')
gfl_lancif=gfl_lancif[['LAN_ID','SUB_CATEGORY','CUSTOMER_RESIDENTIAL_STATUS']]

loan_view=pd.concat([ghf_loan_view,gfl_loan_view]) #REFERENCE
base_view=pd.concat([ghf_base_view,gfl_base_view]) # LAN_ID
# base_view.to_excel(r"C:\Users\VAIBHAV.SRIVASTAV01\Desktop\base_view.xlsx")
cust_view=pd.concat([lancif,gfl_lancif])
loan_view.rename(columns={'REFERENCE':'LAN_ID'},inplace=True)
loan_view=loan_view[['LAN_ID','LOAN_PURPOSE','ROI','PRINCIPAL_OUTSTANDING','GPL_FLAG','SUB_PRODUCT']]
base_view=base_view.merge(loan_view, on='LAN_ID',how='left')
disb=pd.concat([ghf_disb,gfl_disb])
disb=disb[['FINANCE_REFERENCE','DISBURSEMENT_DATE', 'DISBURSEMENT_AMOUNT','DISBURSEMENT_SEQUENCE']]
disb.rename(columns={'FINANCE_REFERENCE':'LAN_ID'},inplace=True)
base_view['FINTYPE']=np.where(base_view['LOAN_PURPOSE'].isin(['LAP Balance Transfer plus Top-up ',' Loan against Property ', 'Industrial LAP Balance Transfer','Industrial LAP Balance Transfer plus Top-up','LAP Balance Transfer','Loan against industrial property','LAP Top Up']),'LP',base_view['FINANCE_TYPE'])
# branch=base_view.copy()
# branch=branch[['LAN_ID','FINTYPE','REPORTING_BRANCH']]
# disb=disb.merge(branch,on='LAN_ID',how='left')
# disb=disb[['FINTYPE','REPORTING_BRANCH','DISBURSEMENT_DATE', 'DISBURSEMENT_AMOUNT']]
base_view['GPLFLAG_SANCTIONS']=np.where((base_view['FINTYPE'].isin(['LP','NP'])==False) & (base_view['GPL_FLAG']=='YES') & (base_view['NBFC_FLAG']=='N') , 'GPL',np.where((base_view['FINTYPE'].isin(['LP','NP'])==False) &(base_view['GPL_FLAG']=='NO') & (base_view['NBFC_FLAG']=='N'),'NON GPL','NIL'))
# dtcron=pd.read_sql("select finreference from dtcron where fs_year_month='2022-12' and status_seg='A) Final Sanction';",con)
ins=pd.concat([ghf_ins,gfl_ins])
base_view=base_view.merge(ins,on='LAN_ID',how='left')
base_view=base_view.merge(cust_view,on='LAN_ID',how='left')
def error_lans():
    a=['GHF1001FL0000328',
 'GHF1001FL0002120',
 'GHF1002FL0002436',
 'GHF1001FL0002119',
 'GHF1002FL0000299',
 'GHF1002FL0000640',
 'GHF1002HL0002943',
 'GHF1002HL0000836',
 'GHF1002FT0001610',
 'GHF1002FL0001993',
 'GHF1002HL0002773',
 'GHF1001FT0000552',
 'GHF1101LP0003918',
 'GHF1401HL0000123',
 'GHF1002HL0003082',
 'GHF1002FL0000605',
 'GHF1002HL0002977',
 'GHF1001FL0000327',
 'GHF1002HL0003004',
 'GHF1002FL0005334',
 'GHF1001FL0000912',
 'GHF1002FL0000257',
 'GHF1002HL0003367',
 'GHF1002FL0000744',
 'GHF1002FT0001009',
 'GHF1001FL0000329',
 'GHF1401HL0000122',
 'GHF1001HL0000342',
 'GHF1002HL0001014',
 'GHF1002HL0003484',
 'GHF1401HL0001902',
 'GHF1001HL0000900',
 'GHF1002FL0000063',
 'GHF1001FL0002118',
 'GHF1401HL0000124',
 'GHF1002FL0002252',
 'GHF1401HL0000125',
 'GHF1002FL0003178',
 'GHF1002FL0000659',
 'GHF1001HL0000967',
 'GHF1001HL0000966']
    return a
def remove_error_lans(df):
    try:
        df=df[df['REFERENCE'].isin(error_lans())==False]
    except:
        pass
    try:
        df=df[df['LAN_ID'].isin(error_lans())==False]
    except:
        pass
    try:
        df=df[df['FINREFERENCE'].isin(error_lans())==False]
    except:
        pass
    try:
        df=df[df['FINANCE_REFERENCE'].isin(error_lans())==False]
    except:
        pass
    return df
base_view=remove_error_lans(base_view)

eomfin=base_view.copy()
eomfin['EOMLOGN'].replace(to_replace=[np.nan,'',' '],value=date.date(1900,1,1),inplace=True)
eomfin['EOMSNCTN'].replace(to_replace=[np.nan,'',' '],value=date.date(1900,1,1),inplace=True)
eomfin['EOMRJCT'].replace(to_replace=[np.nan,'',' '],value=date.date(1900,1,1),inplace=True)
eomfin['EOMCLTRL'].replace(to_replace=[np.nan,'',' '],value=date.date(1900,1,1),inplace=True)
eomfin['BOOKING_DATE'].replace(to_replace=[np.nan,'',' '],value=date.date(1900,1,1),inplace=True)



eomfin['REPORTING_BRANCH']=np.where(eomfin['REPORTING_BRANCH']=='Aundh BO (Pune)','Pune',eomfin['REPORTING_BRANCH'])
eomfin['REPORTING_BRANCH']=np.where(eomfin['REPORTING_BRANCH']=='Bangalore (MG Road)','Bangalore',eomfin['REPORTING_BRANCH'])
eomfin['REPORTING_BRANCH']=np.where(eomfin['REPORTING_BRANCH']=='Banglore (Sahakar Nagar)','Bangalore',eomfin['REPORTING_BRANCH'])
eomfin['REPORTING_BRANCH']=np.where(eomfin['REPORTING_BRANCH'].isin(['Gurgaon','Gautam Buddha Nagar']),'Delhi',eomfin['REPORTING_BRANCH'])
eomfin['REPORTING_BRANCH']=np.where(eomfin['REPORTING_BRANCH']=='Thane','Mumbai',eomfin['REPORTING_BRANCH'])

bussdate=pd.read_sql("select * from prod_bi.aabi.business_date;",con_dm)
bussdate1=bussdate.copy()
today=date.datetime.now()

bussdate1.rename(columns={'BUSINESS_DATE':'EOMLOGN'},inplace=True)
eomfin.rename(columns={'EOMLOGN':'CAL_DATE'},inplace=True)
eomfin['CAL_DATE']=pd.to_datetime(eomfin['CAL_DATE']).dt.date
eomfin=eomfin.merge(bussdate1,on='CAL_DATE')
# eomfin['EOMLOGN']=np.where(pd.to_datetime(eomfin['CAL_DATE']).dt.date>pd.to_datetime(eomfin['BOOKING_DATE']).dt.date,pd.to_datetime(eomfin['BOOKING_DATE']).dt.date,(pd.to_datetime(eomfin['EOMLOGN']).dt.date))
# eomfin['EOMSNCTN']=np.where(pd.to_datetime(eomfin['EOMSNCTN']).dt.date>pd.to_datetime(eomfin['BOOKING_DATE']).dt.date,pd.to_datetime(eomfin['BOOKING_DATE']).dt.date,(pd.to_datetime(eomfin['EOMSNCTN']).dt.date))

bussdate2=bussdate.copy()
bussdate2.rename(columns={'BUSINESS_DATE':'EOMSNCTN'},inplace=True)
eomfin.rename(columns={'EOMSNCTN':'PSV1'},inplace=True)
bussdate2.rename(columns={'CAL_DATE':'PSV1'},inplace=True)
eomfin['PSV1']=pd.to_datetime(eomfin['PSV1']).dt.date
eomfin['PSV1'].replace(to_replace=[np.nan,'',' '],value=date.date(1900,1,1),inplace=True)
eomfin=eomfin.merge(bussdate2,on='PSV1',how='left')
eomfin['EOMSNCTN'].replace(to_replace=[np.nan,'',' '],value=date.date(1900,1,1),inplace=True)


bussdate3=bussdate.copy()
bussdate3.rename(columns={'BUSINESS_DATE':'BOOKING_DATE'},inplace=True)
eomfin.rename(columns={'BOOKING_DATE':'booked_date'},inplace=True)
bussdate3.rename(columns={'CAL_DATE':'booked_date'},inplace=True)
eomfin['booked_date']=pd.to_datetime(eomfin['booked_date']).dt.date
eomfin['booked_date'].replace(to_replace=[np.nan,'',' '],value=date.date(1900,1,1),inplace=True)
eomfin=eomfin.merge(bussdate3,on='booked_date',how='left')
eomfin['BOOKING_DATE'].replace(to_replace=[np.nan,'',' '],value=date.date(1900,1,1),inplace=True)

bussdate4=bussdate.copy()
bussdate4.rename(columns={'BUSINESS_DATE':'EOMRJCT'},inplace=True)
eomfin.rename(columns={'EOMRJCT':'REJECT_DATE'},inplace=True)
bussdate4.rename(columns={'CAL_DATE':'REJECT_DATE'},inplace=True)
eomfin['REJECT_DATE']=pd.to_datetime(eomfin['REJECT_DATE']).dt.date
eomfin['REJECT_DATE'].replace(to_replace=[np.nan,'',' '],value=date.date(1900,1,1),inplace=True)
eomfin=eomfin.merge(bussdate4,on='REJECT_DATE',how='left')
eomfin['EOMRJCT'].replace(to_replace=[np.nan,'',' '],value=date.date(1900,1,1),inplace=True)

bussdate5=bussdate.copy()
bussdate5.rename(columns={'BUSINESS_DATE':'EOMCLTRL'},inplace=True)
eomfin.rename(columns={'EOMCLTRL':'IS_DATE'},inplace=True)
bussdate5.rename(columns={'CAL_DATE':'IS_DATE'},inplace=True)
eomfin['IS_DATE']=pd.to_datetime(eomfin['IS_DATE']).dt.date
eomfin['IS_DATE'].replace(to_replace=[np.nan,'',' '],value=date.date(1900,1,1),inplace=True)
eomfin=eomfin.merge(bussdate5,on='IS_DATE',how='left')
eomfin['EOMCLTRL'].replace(to_replace=[np.nan,'',' '],value=date.date(1900,1,1),inplace=True)


eomfin['LOGIN_YEAR']=pd.to_datetime(eomfin['EOMLOGN']).dt.year
eomfin['LOGIN_YEAR']=eomfin['LOGIN_YEAR'].astype(int)
eomfin['LOGIN_MONTH']=pd.to_datetime(eomfin['EOMLOGN']).dt.month
eomfin['LOGIN_MONTH']=eomfin['LOGIN_MONTH'].astype(int)
eomfin['FS_YEAR']=pd.to_datetime(eomfin['EOMSNCTN']).dt.year
eomfin['FS_YEAR']=eomfin['FS_YEAR'].astype(int)
eomfin['FS_MONTH']=pd.to_datetime(eomfin['EOMSNCTN']).dt.month
eomfin['FS_MONTH']=eomfin['FS_MONTH'].astype(int)
eomfin['BOOKING_YEAR']=pd.to_datetime(eomfin['BOOKING_DATE']).dt.year
eomfin['BOOKING_YEAR']=eomfin['BOOKING_YEAR'].astype(int)
eomfin['BOOKING_MONTH']=pd.to_datetime(eomfin['BOOKING_DATE']).dt.month
eomfin['BOOKING_MONTH']=eomfin['BOOKING_MONTH'].astype(int)
eomfin['REJECT_YEAR']=pd.to_datetime(eomfin['EOMRJCT']).dt.year
eomfin['REJECT_YEAR']=eomfin['REJECT_YEAR'].astype(int)
eomfin['REJECT_MONTH']=pd.to_datetime(eomfin['EOMRJCT']).dt.month
eomfin['REJECT_MONTH']=eomfin['REJECT_MONTH'].astype(int)
eomfin['IS_YEAR']=pd.to_datetime(eomfin['EOMCLTRL']).dt.year
eomfin['IS_YEAR']=eomfin['IS_YEAR'].astype(int)
eomfin['IS_MONTH']=pd.to_datetime(eomfin['EOMCLTRL']).dt.month
eomfin['IS_MONTH']=eomfin['IS_MONTH'].astype(int)
disb['DISBURSE_MONTH']=pd.to_datetime(disb['DISBURSEMENT_DATE']).dt.month
disb['DISBURSE_YEAR']=pd.to_datetime(disb['DISBURSEMENT_DATE']).dt.year
eomfin['LOGIN_YEAR_MONTH']=eomfin['LOGIN_YEAR'].astype(str)+'-'+ eomfin['LOGIN_MONTH'].astype(str)
eomfin['FS_YEAR_MONTH']=eomfin['FS_YEAR'].astype(str)+'-'+ eomfin['FS_MONTH'].astype(str)
eomfin['BOOK_YEAR_MONTH']=eomfin['BOOKING_YEAR'].astype(str)+'-'+ eomfin['BOOKING_MONTH'].astype(str)
disb['DISB_YEAR_MONTH']=disb['DISBURSE_YEAR'].astype(str)+'-'+ disb['DISBURSE_MONTH'].astype(str)
eomfin['REJECT_YEAR_MONTH']=eomfin['REJECT_YEAR'].astype(str)+'-'+ eomfin['REJECT_MONTH'].astype(str)
eomfin['IS_YEAR_MONTH']=eomfin['IS_YEAR'].astype(str)+'-'+ eomfin['IS_MONTH'].astype(str)
eomfin['BOOKING_DATE']=np.where(eomfin['LAN_ID'].isin(['GHF1002HL0010862' ,'GHF1002FL0010865' ,'GHF1002HL0010527' ,'GHF1002HL0010885' ,'GHF1002FL0010994' ,'GHF1003HL0010367' ,'GHF1003FL0010946' ,'GHF1002HL0010399' ,'GHF1002HT0010434' ,'GHF1003HL0009576' ,'GHF1002HL0010722' ,'GHF1002HL0010827']),date.date(2022,4,1),pd.to_datetime(eomfin['BOOKING_DATE']).dt.date)
eomfin['BOOK_YEAR_MONTH']=np.where(eomfin['LAN_ID'].isin(['GHF1002HL0010862' ,'GHF1002FL0010865' ,'GHF1002HL0010527' ,'GHF1002HL0010885' ,'GHF1002FL0010994' ,'GHF1003HL0010367' ,'GHF1003FL0010946' ,'GHF1002HL0010399' ,'GHF1002HT0010434' ,'GHF1003HL0009576' ,'GHF1002HL0010722' ,'GHF1002HL0010827']),'2022-4',eomfin['BOOK_YEAR_MONTH'])
eomfin=eomfin[['REPORTING_BRANCH','LAN_ID',  'FINTYPE', 'LOAN_PURPOSE', 'LOAN_STATUS',
       'BOOKING_AMOUNT', 'BOOKING_DATE', 'ROI', 'NET_PREMIUM',
       'FINANCE_SOURCE_ID', 'EOMLOGN', 'EOMSNCTN', 'REQUESTED_AMOUNT',
       'SANCTION_AMOUNT', 'NBFC_FLAG','SUB_CATEGORY','CUSTOMER_RESIDENTIAL_STATUS',
       'DETAILED_STATUS', 'STATUS', 'QUEUE', 'LOGIN_STATUS', 'STATUS_SEG',
       'LOGIN_MONTH', 'LOGIN_YEAR_MONTH', 'FS_MONTH', 'FS_YEAR_MONTH',
       'BOOKING_MONTH', 'BOOK_YEAR_MONTH','REJECT_YEAR_MONTH','IS_YEAR_MONTH','EOMRJCT','EOMCLTRL','GPLFLAG_SANCTIONS','SUB_PRODUCT','INS_FLAG']]
eomdisb_1=disb.copy()


###############################################################################
'''
RUN CHECKER
'''
data_length=len(eomfin)
duplicate_df=eomfin[eomfin.duplicated(['LAN_ID'])]
if len(duplicate_df)!=0:
    duplicate_flag='Y'
else:
    duplicate_flag='N'
    
import logging
import datetime as datetime
from datetime import timedelta
import time
import os
yesterday = datetime.datetime.now() - timedelta(1)
current_dir="//GHFL-SNOWFLAKES/Users/bischeduler/Documents/Output_of_Scheduler/BI_Daily_Reports"
os.chdir(f'{current_dir}/{yesterday.strftime("/%Y-%m-%d")}')
logging.basicConfig(filename="GC1" +yesterday.strftime("%Y-%m-%d")+".log",format='%(asctime)s - %(message)s', level=logging.INFO)
logging.basicConfig(filename="GC1" +yesterday.strftime("%Y-%m-%d")+".log" , filemode='w', format='%(name)s - %(levelname)s - %(message)s')
logging.info("NUMBER OF ROWS = " +str(data_length) + "; DUPLICATIONS = "+duplicate_flag)


if duplicate_flag =='N':

    
    ################################ SUMMARY ################################
    def end_of(MTD,eomfin,eomdisb_1,lastmonth=None,LRD=None,Rd=None,LAP='LAP'):
        eomfin_book=eomfin.copy()
        if type(MTD)==datetime.datetime:
            eomfin_book['bookdate']=np.where(True,pd.to_datetime(eomfin_book['BOOKING_DATE']).dt.date ,0)
            eomfin_book=eomfin_book[eomfin_book['bookdate']==(MTD.date())]
            # eomfin_book['BOOK_VOL']= np.where(eomfin_book['bookdate']==(MTD.date()),1,0)
        elif lastmonth!=None:
            eomfin_book['bookdate']=np.where(True,pd.to_datetime(eomfin_book['BOOKING_DATE']).dt.date ,0)
            eomfin_book=eomfin_book[(eomfin_book['bookdate']<=(lastmonth.date())) & (eomfin_book['BOOK_YEAR_MONTH'].isin(MTD))]
        else:
            eomfin_book=eomfin_book[eomfin_book['BOOK_YEAR_MONTH'].isin(MTD)]
        
            # eomfin_book['BOOK_VOL']= np.where((eomfin_book['BOOK_YEAR_MONTH'].isin(MTD)),1,0)
        eomfin_book=eomfin_book[eomfin_book['STATUS']=='Booked']
        eomfin_book['BOOK_VOL']= np.where(eomfin_book['STATUS']=='Booked',1,0)
        
        eomfin_book['BOOKING_AMOUNT']=eomfin_book['BOOKING_AMOUNT'].round(2)
        eomfin_book['WROI']=eomfin_book['ROI']*eomfin_book['BOOKING_AMOUNT']
        eomfin_book=eomfin_book[[ 'LAN_ID','REPORTING_BRANCH', 'BOOKING_AMOUNT','BOOK_VOL','FINTYPE','WROI','NET_PREMIUM','GPLFLAG_SANCTIONS','SUB_PRODUCT']]
        eomfin_bookview=eomfin_book.groupby('REPORTING_BRANCH').sum(['BOOKING_AMOUNT','BOOK_VOL','WROI'])
        eomfin_bookview['ROI']=eomfin_bookview['WROI']/eomfin_bookview['BOOKING_AMOUNT']
        eomfin_bookview['Premium']=eomfin_bookview['NET_PREMIUM']/10000000
        eomfin_bookview['Penetration']=eomfin_bookview['NET_PREMIUM']/eomfin_bookview['BOOKING_AMOUNT']
        eomfin_bookview['BOOKING_AMOUNT']=eomfin_bookview['BOOKING_AMOUNT']/10000000
        eomfin_bookview=eomfin_bookview[[  'BOOKING_AMOUNT','BOOK_VOL','ROI','Premium','Penetration']]
        eomfin_book['Prod']=np.where(((eomfin_book['FINTYPE'].isin(['LP','NP']) & (eomfin_book['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP',LRD])==False))),LAP,np.where((eomfin_book['FINTYPE'].isin(['LP','NP']) & (eomfin_book['SUB_PRODUCT'].isin([LRD]))),LRD,np.where((eomfin_book['FINTYPE'].isin(['LP','NP']) & (eomfin_book['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP']))),'NEO-LAP','HL')))
        eomfin_pbview=eomfin_book.groupby('Prod').sum(['BOOKING_AMOUNT','BOOK_VOL','WROI'])
        eomfin_pbview['ROI']=eomfin_pbview['WROI']/eomfin_pbview['BOOKING_AMOUNT']
        eomfin_pbview['Premium']=eomfin_pbview['NET_PREMIUM']/10000000
        eomfin_pbview['Penetration']=eomfin_pbview['NET_PREMIUM']/eomfin_pbview['BOOKING_AMOUNT']
        eomfin_pbview['BOOKING_AMOUNT']=eomfin_pbview['BOOKING_AMOUNT']/10000000
        eomfin_pbview=eomfin_pbview[[  'BOOKING_AMOUNT','BOOK_VOL','ROI','Premium','Penetration']]
        
        eomfin_log=eomfin.copy()
        
        if type(MTD)==datetime.datetime:
            eomfin_log['logdate']=np.where(True,pd.to_datetime(eomfin_log['EOMLOGN']).dt.date ,0)
            eomfin_log=eomfin_log[eomfin_log['logdate']==(MTD.date())]
            # eomfin_log['LOG_VOL']= np.where(eomfin_log['logdate']==(MTD.date()),1,0)
        elif lastmonth!=None:
            eomfin_log['logdate']=np.where(True,pd.to_datetime(eomfin_log['EOMLOGN']).dt.date ,0)
            eomfin_log=eomfin_log[(eomfin_log['logdate']<=(lastmonth.date())) & (eomfin_log['LOGIN_YEAR_MONTH'].isin(MTD))]
        else:
            eomfin_log=eomfin_log[eomfin_log['LOGIN_YEAR_MONTH'].isin(MTD)]
            # eomfin_log['LOG_VOL']= np.where((eomfin_log['LOGIN_YEAR_MONTH'].isin(MTD)),1,0)
        eomfin_log['LOG_VOL']= np.where(eomfin_log['LOGIN_STATUS']=='A) Login',1,0)
        eomfin_log=eomfin_log[eomfin_log['LOGIN_STATUS']=='A) Login']
        eomfin_log['REQUESTED_AMOUNT']=eomfin_log['REQUESTED_AMOUNT']/10000000
        eomfin_log['REQUESTED_AMOUNT']=eomfin_log['REQUESTED_AMOUNT'].round(2)
        eomfin_log=eomfin_log[[ 'LAN_ID', 'REPORTING_BRANCH','REQUESTED_AMOUNT','LOG_VOL','FINTYPE','GPLFLAG_SANCTIONS','SUB_PRODUCT']]
        eomfin_logview=eomfin_log.groupby('REPORTING_BRANCH').sum(['REQUESTED_AMOUNT','LOG_VOL'])
        eomfin_log['Prod']=np.where(((eomfin_log['FINTYPE'].isin(['LP','NP']) & (eomfin_log['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP',LRD])==False))),LAP,np.where((eomfin_log['FINTYPE'].isin(['LP','NP']) & (eomfin_log['SUB_PRODUCT'].isin([LRD]))),LRD,np.where((eomfin_log['FINTYPE'].isin(['LP','NP']) & (eomfin_log['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP']))),'NEO-LAP','HL'))) 
        eomfin_plview=eomfin_log.groupby('Prod').sum(['REQUESTED_AMOUNT','LOG_VOL'])
        
        
        eomfin_sanc=eomfin.copy()
        
        if type(MTD)==datetime.datetime:
            eomfin_sanc['sancdate']=np.where(True,pd.to_datetime(eomfin_sanc['EOMSNCTN']).dt.date ,0)
            eomfin_sanc=eomfin_sanc[eomfin_sanc['sancdate']==(MTD.date())]
            # eomfin_sanc['SANCTION_VOL']= np.where((eomfin_sanc['sancdate']==(MTD.date())),1,0)
        elif lastmonth!=None:
            eomfin_sanc['sancdate']=np.where(True,pd.to_datetime(eomfin_sanc['EOMSNCTN']).dt.date ,0)
            eomfin_sanc=eomfin_sanc[(eomfin_sanc['sancdate']<=(lastmonth.date())) & (eomfin_sanc['FS_YEAR_MONTH'].isin(MTD))]
        else:
            eomfin_sanc=eomfin_sanc[eomfin_sanc['FS_YEAR_MONTH'].isin(MTD)]
            # eomfin_sanc['SANCTION_VOL']= np.where((eomfin_sanc['FS_YEAR_MONTH'].isin(MTD)),1,0)
        eomfin_sanc['SANCTION_VOL']= np.where(eomfin_sanc['STATUS_SEG']=='A) Final Sanction',1,0)
        eomfin_sanc=eomfin_sanc[eomfin_sanc['STATUS_SEG']=='A) Final Sanction']
        eomfin_sanc['SANCTION_AMOUNT']=eomfin_sanc['SANCTION_AMOUNT']/10000000
        eomfin_sanc['SANCTION_AMOUNT']=eomfin_sanc['SANCTION_AMOUNT'].round(2)
        eomfin_sanc=eomfin_sanc[[ 'LAN_ID','REPORTING_BRANCH', 'SANCTION_AMOUNT','SANCTION_VOL','FINTYPE','GPLFLAG_SANCTIONS','SUB_PRODUCT']]
        eomfin_sancview=eomfin_sanc.groupby('REPORTING_BRANCH').sum(['SANCTION_AMOUNT','SANCTION_VOL'])
    
        eomfin_sanc['Prod']=np.where(((eomfin_sanc['FINTYPE'].isin(['LP','NP']) & (eomfin_sanc['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP',LRD])==False))),LAP,np.where((eomfin_sanc['FINTYPE'].isin(['LP','NP']) & (eomfin_sanc['SUB_PRODUCT'].isin([LRD]))),LRD,np.where((eomfin_sanc['FINTYPE'].isin(['LP','NP']) & (eomfin_sanc['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP']))),'NEO-LAP','HL')))
        eomfin_psview=eomfin_sanc.groupby('Prod').sum(['SANCTION_AMOUNT','SANCTION_VOL'])
        
        eomfin_insanc=eomfin.copy()
        
        if type(MTD)==datetime.datetime:
            eomfin_insanc['inc_sancdate']=np.where(True,pd.to_datetime(eomfin_insanc['EOMCLTRL']).dt.date ,0)
            eomfin_insanc=eomfin_insanc[eomfin_insanc['inc_sancdate']==(MTD.date())]
            # eomfin_insanc['SANCTION_VOL']= np.where((eomfin_insanc['sancdate']==(MTD.date())),1,0)
        elif lastmonth!=None:
            eomfin_insanc['inc_sancdate']=np.where(True,pd.to_datetime(eomfin_insanc['EOMCLTRL']).dt.date ,0)
            eomfin_insanc=eomfin_insanc[(eomfin_insanc['inc_sancdate']<=(lastmonth.date())) & (eomfin_insanc['IS_YEAR_MONTH'].isin(MTD))]
        else:
            eomfin_insanc=eomfin_insanc[eomfin_insanc['IS_YEAR_MONTH'].isin(MTD)]
            # eomfin_insanc['SANCTION_VOL']= np.where((eomfin_insanc['FS_YEAR_MONTH'].isin(MTD)),1,0)
        eomfin_insanc['IN_SANCTION_VOL']= np.where(eomfin_insanc['STATUS_SEG']=='C) Income Sanction',1,0)
        eomfin_insanc=eomfin_insanc[eomfin_insanc['STATUS_SEG']=='C) Income Sanction']
        eomfin_insanc['IN_SANCTION_AMOUNT']=eomfin_insanc['SANCTION_AMOUNT']/10000000
        eomfin_insanc['IN_SANCTION_AMOUNT']=eomfin_insanc['IN_SANCTION_AMOUNT'].round(2)
        eomfin_insanc=eomfin_insanc[[ 'LAN_ID','REPORTING_BRANCH', 'IN_SANCTION_AMOUNT','IN_SANCTION_VOL','FINTYPE','GPLFLAG_SANCTIONS','SUB_PRODUCT']]
        eomfin_insancview=eomfin_insanc.groupby('REPORTING_BRANCH').sum(['IN_SANCTION_AMOUNT','IN_SANCTION_VOL'])
        eomfin_insanc['Prod']=np.where(((eomfin_insanc['FINTYPE'].isin(['LP','NP']) & (eomfin_insanc['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP',LRD])==False))),LAP,np.where((eomfin_insanc['FINTYPE'].isin(['LP','NP']) & (eomfin_insanc['SUB_PRODUCT'].isin([LRD]))),LRD,np.where((eomfin_insanc['FINTYPE'].isin(['LP','NP']) & (eomfin_insanc['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP']))),'NEO-LAP','HL')))
        eomfin_inpsview=eomfin_insanc.groupby('Prod').sum(['IN_SANCTION_AMOUNT','IN_SANCTION_VOL'])
        
        eomfin_reject=eomfin.copy()
        
        if type(MTD)==datetime.datetime:
            eomfin_reject['rejectdate']=np.where(True,pd.to_datetime(eomfin_reject['EOMRJCT']).dt.date ,0)
            eomfin_reject=eomfin_reject[eomfin_reject['rejectdate']==(MTD.date())]
            # eomfin_reject['SANCTION_VOL']= np.where((eomfin_reject['sancdate']==(MTD.date())),1,0)
        elif lastmonth!=None:
            eomfin_reject['rejectdate']=np.where(True,pd.to_datetime(eomfin_reject['EOMRJCT']).dt.date ,0)
            eomfin_reject=eomfin_reject[(eomfin_reject['rejectdate']<=(lastmonth.date())) & (eomfin_reject['REJECT_YEAR_MONTH'].isin(MTD))]
        else:
            eomfin_reject=eomfin_reject[eomfin_reject['REJECT_YEAR_MONTH'].isin(MTD)]
            # eomfin_reject['SANCTION_VOL']= np.where((eomfin_reject['FS_YEAR_MONTH'].isin(MTD)),1,0)
        eomfin_reject['REJECT_VOL']= np.where(eomfin_reject['STATUS_SEG']=='B) Rejected',1,0)
        eomfin_reject=eomfin_reject[eomfin_reject['STATUS_SEG']=='B) Rejected']
        eomfin_reject['REJECT_AMOUNT']=eomfin_reject['REQUESTED_AMOUNT']/10000000
        eomfin_reject['REJECT_AMOUNT']=eomfin_reject['REJECT_AMOUNT'].round(2)
        eomfin_reject=eomfin_reject[[ 'LAN_ID','REPORTING_BRANCH', 'REJECT_AMOUNT','REJECT_VOL','FINTYPE','GPLFLAG_SANCTIONS','SUB_PRODUCT']]
        eomfin_rejectview=eomfin_reject.groupby('REPORTING_BRANCH').sum(['REJECT_AMOUNT','REJECT_VOL'])
        eomfin_reject['Prod']=np.where(((eomfin_reject['FINTYPE'].isin(['LP','NP']) & (eomfin_reject['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP',LRD])==False))),LAP,np.where((eomfin_reject['FINTYPE'].isin(['LP','NP']) & (eomfin_reject['SUB_PRODUCT'].isin([LRD]))),LRD,np.where((eomfin_reject['FINTYPE'].isin(['LP','NP']) & (eomfin_reject['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP']))),'NEO-LAP','HL')))     
        
        eomfin_rejectpsview=eomfin_reject.groupby('Prod').sum(['REJECT_AMOUNT','REJECT_VOL'])
        
        
        
        m1=pd.merge(eomfin_bookview,eomfin_logview,left_on=('REPORTING_BRANCH'),right_on=('REPORTING_BRANCH'), how='outer')
        m2=pd.merge(m1,eomfin_sancview,left_on=('REPORTING_BRANCH'),right_on=('REPORTING_BRANCH'), how='outer')
        m3=pd.merge(m2,eomfin_insancview,left_on=('REPORTING_BRANCH'),right_on=('REPORTING_BRANCH'), how='outer')
        eomfinview=pd.merge(m3,eomfin_rejectview,left_on=('REPORTING_BRANCH'),right_on=('REPORTING_BRANCH'), how='outer')
        
        
        t1=pd.merge(eomfin_pbview,eomfin_plview,left_on=('Prod'),right_on=('Prod'), how='outer')
        t2=pd.merge(t1,eomfin_psview,left_on=('Prod'),right_on=('Prod'), how='outer')
        t3=pd.merge(t2,eomfin_inpsview,left_on=('Prod'),right_on=('Prod'), how='outer')
        eompfinview=pd.merge(t3,eomfin_rejectpsview,left_on=('Prod'),right_on=('Prod'), how='outer')
        
        eomdisb=pd.merge(eomfin,eomdisb_1,on=('LAN_ID'),how='left')
        # eomdisb=eomdisb_1.copy()
        if type(MTD)==datetime.datetime:
            eomdisb['DISBDATE']=np.where(True,pd.to_datetime(eomdisb['DISBURSEMENT_DATE']).dt.date ,date.date(1900,1,1))
            eomdisb=eomdisb[eomdisb['DISBDATE']==(MTD.date())]
            # eomdisb['DISB_VOL']=np.where((eomdisb['DISBDATE']==(MTD.date())),1,0)
        elif lastmonth!=None:
            eomdisb['DISBDATE']=np.where(True,pd.to_datetime(eomdisb['DISBURSEMENT_DATE']).dt.date ,date.date(1900,1,1))
            eomdisb=eomdisb[(eomdisb['DISBDATE']<=(lastmonth.date())) & (eomdisb['DISB_YEAR_MONTH'].isin(MTD))]
        else:
            eomdisb=eomdisb[eomdisb['DISB_YEAR_MONTH'].isin(MTD)]
            # eomdisb['DISB_VOL']=np.where((eomdisb['DISB_YEAR_MONTH'].isin(MTD)),1,0)
        eomdisb['DISB_VOL']=np.where(True,1,0)
        eomdisb['DISBURSEMENT_AMOUNT_TRANCH1']=np.where(eomdisb['DISBURSEMENT_SEQUENCE']==1,eomdisb['DISBURSEMENT_AMOUNT']/1000000000,0)
        eomdisb['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']=np.where(eomdisb['DISBURSEMENT_SEQUENCE']!=1,eomdisb['DISBURSEMENT_AMOUNT']/1000000000,0)
        # eomdisb['DISBURSEMENT_AMOUNT_TRANCH1']=eomdisb['DISBURSEMENT_AMOUNT_TRANCH1'].round(2)
        # eomdisb['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']=eomdisb['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES'].round(2)
        eomdisb=eomdisb[[ 'REPORTING_BRANCH', 'DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES','FINTYPE','GPLFLAG_SANCTIONS','SUB_PRODUCT']]
        eomdisbview=eomdisb.groupby('REPORTING_BRANCH').sum(['DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES'])
        eomdisb['Prod']=np.where(((eomdisb['FINTYPE'].isin(['LP','NP']) & (eomdisb['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP',LRD])==False))),LAP,np.where((eomdisb['FINTYPE'].isin(['LP','NP']) & (eomdisb['SUB_PRODUCT'].isin([LRD]))),LRD,np.where((eomdisb['FINTYPE'].isin(['LP','NP']) & (eomdisb['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP']))),'NEO-LAP','HL')))     
        eomfin_pdview=eomdisb.groupby('Prod').sum(['DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES'])
        
        eomp=pd.merge(eompfinview,eomfin_pdview,left_on=('Prod'),right_on=('Prod'), how='outer')
        eom=pd.merge(eomfinview,eomdisbview,left_on=('REPORTING_BRANCH'),right_on=('REPORTING_BRANCH'), how='outer')
        eom=eom.sort_values(by='REPORTING_BRANCH')
        eomp=eomp.sort_values(by='Prod')
        eom=eom[['LOG_VOL','REQUESTED_AMOUNT','SANCTION_VOL','SANCTION_AMOUNT','BOOK_VOL','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES','REJECT_AMOUNT','REJECT_VOL','IN_SANCTION_AMOUNT','IN_SANCTION_VOL','ROI','Premium','Penetration']]
        eomp=eomp[['LOG_VOL','REQUESTED_AMOUNT','SANCTION_VOL','SANCTION_AMOUNT','BOOK_VOL','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES','REJECT_AMOUNT','REJECT_VOL','IN_SANCTION_AMOUNT','IN_SANCTION_VOL','ROI','Premium','Penetration']]
        
        monbr=list(set(list(base_view.REPORTING_BRANCH.unique()))-set(eom.index.to_list()))
        if LRD==None:
            missprod=list(set(['HL','NEO-LAP',LAP])-set(eomp.index.to_list()))
        else:
            missprod=list(set(['HL','NEO-LAP',LAP,LRD])-set(eomp.index.to_list()))
        for i in monbr:
            eom.loc[i,:]=0
        for i in missprod:
            eomp.loc[i,:]=0
        eom=eom.sort_values(by='REPORTING_BRANCH')
        eomp=eomp.sort_values(by='Prod')
        eom.replace(to_replace=[np.nan,'',' '],value=0,inplace=True)
        eomp.replace(to_replace=[np.nan,'',' '],value=0,inplace=True)
        a=eom['ROI']*eom['BOOKING_AMOUNT']
        a=pd.DataFrame(a)
        a['BOOKING_AMOUNT']=eom['BOOKING_AMOUNT']
        a['NET_PREMIUM']=eom['Premium']
        a.loc['Total']= a.sum()
        eom.loc['Total']= eom.sum()
        eom.loc['Total','ROI']=a.loc['Total',0]/a.loc['Total','BOOKING_AMOUNT']
        eom.loc['Total','Penetration']=a.loc['Total','NET_PREMIUM']/a.loc['Total','BOOKING_AMOUNT']
        eomp.loc['Total']= eomp.sum()
        eomp.loc['Total','ROI']=a.loc['Total',0]/a.loc['Total','BOOKING_AMOUNT']
        eomp.loc['Total','Penetration']=a.loc['Total','NET_PREMIUM']/a.loc['Total','BOOKING_AMOUNT']
        eom['Disbursement_Volume']=eom['BOOK_VOL']
        eomp['Disbursement_Volume']=eomp['BOOK_VOL']
        
        eom=eom[['LOG_VOL','REQUESTED_AMOUNT','SANCTION_VOL','SANCTION_AMOUNT','BOOK_VOL','BOOKING_AMOUNT','Disbursement_Volume','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES','REJECT_AMOUNT','REJECT_VOL','IN_SANCTION_AMOUNT','IN_SANCTION_VOL','ROI','Premium','Penetration']]
        eomp=eomp[['LOG_VOL','REQUESTED_AMOUNT','SANCTION_VOL','SANCTION_AMOUNT','BOOK_VOL','BOOKING_AMOUNT','Disbursement_Volume','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES','REJECT_AMOUNT','REJECT_VOL','IN_SANCTION_AMOUNT','IN_SANCTION_VOL','ROI','Premium','Penetration']]
        eom.replace(to_replace=[np.nan,'',' '],value=0,inplace=True)
        eomp.replace(to_replace=[np.nan,'',' '],value=0,inplace=True)
        if Rd==True:
            eom=eom.round(2)
            eomp=eomp.round(2)
        return eom,eomp
    
    
    
    ##############################################################################
    import datetime as datetime
    from datetime import timedelta
    YTD=[]
    cur_month=todays_date.month
    if cur_month in [1,2,3]:
        for i in [1,2,3]:
            YTD.append(i)
            
    while cur_month>=4:
        YTD.append(cur_month)
        cur_month=cur_month-1
    ytd=[]
    if 1 in YTD:
        for i in YTD:
            if i in [1,2,3]:
                ytd.append(str(todays_date.year)+"-"+str(i))
            else:
                ytd.append(str(todays_date.year-1)+"-"+str(i))
    else:
        for i in YTD:
            ytd.append(str(todays_date.year)+"-"+str(i))
    MTD=[(str(todays_date.year)+"-"+str(todays_date.month))]    
    cal_day=datetime.datetime.now()-timedelta(1)
    bussdate.set_index('CAL_DATE',inplace=True)
    day=bussdate.loc[cal_day.date(),'BUSINESS_DATE']
    MTD_1=[(str(todays_date.year)+"-"+str(todays_date.month-1))]  
    day_1=datetime.datetime(day.year,day.month-1,day.day,0,0)
    day=datetime.datetime(day.year,day.month,day.day,0,0,0,0)
    eom,eomp=end_of(MTD,eomfin,eomdisb_1,LRD='LRD',LAP='Prime')
    eoy,eoyp=end_of(ytd,eomfin,eomdisb_1,LRD='LRD',LAP='Prime')
    eod,eodp=end_of(day,eomfin,eomdisb_1,LRD='LRD',LAP='Prime')
    stlm ,stlmp=end_of(MTD_1,eomfin,eomdisb_1,day_1,LRD='LRD',LAP='Prime')
    mtd_smry=eom[['LOG_VOL', 'REQUESTED_AMOUNT', 'SANCTION_VOL', 'SANCTION_AMOUNT','BOOK_VOL', 'BOOKING_AMOUNT', 'Disbursement_Volume','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','Penetration','ROI']]
    mtdp_smry=eomp[['LOG_VOL', 'REQUESTED_AMOUNT', 'SANCTION_VOL', 'SANCTION_AMOUNT','BOOK_VOL', 'BOOKING_AMOUNT', 'Disbursement_Volume','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES','Premium','Penetration','ROI']]
    ytd_smry=eoy[['LOG_VOL', 'REQUESTED_AMOUNT', 'SANCTION_VOL', 'SANCTION_AMOUNT','BOOK_VOL', 'BOOKING_AMOUNT', 'Disbursement_Volume','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','Penetration','ROI']]
    ytdp_smry=eoyp[['LOG_VOL', 'REQUESTED_AMOUNT', 'SANCTION_VOL', 'SANCTION_AMOUNT','BOOK_VOL', 'BOOKING_AMOUNT', 'Disbursement_Volume','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES','Premium','Penetration','ROI']]
    ftd_smry=eod[['LOG_VOL', 'REQUESTED_AMOUNT', 'SANCTION_VOL', 'SANCTION_AMOUNT','BOOK_VOL', 'BOOKING_AMOUNT', 'Disbursement_Volume','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','Penetration','ROI']]
    ftdp_smry=eodp[['LOG_VOL', 'REQUESTED_AMOUNT', 'SANCTION_VOL', 'SANCTION_AMOUNT','BOOK_VOL', 'BOOKING_AMOUNT', 'Disbursement_Volume','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES','Premium','Penetration','ROI']]
    stlm =stlm[['LOG_VOL', 'REQUESTED_AMOUNT', 'SANCTION_VOL', 'SANCTION_AMOUNT','BOOK_VOL', 'BOOKING_AMOUNT', 'Disbursement_Volume','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','Penetration','ROI']]
    stlmp=stlmp[['LOG_VOL', 'REQUESTED_AMOUNT', 'SANCTION_VOL', 'SANCTION_AMOUNT','BOOK_VOL', 'BOOKING_AMOUNT', 'Disbursement_Volume','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','Penetration','ROI']]
    
    m_prod_br={}
    y_prod_br={}
    d_prod_br={}
    stlm_prod_br={}
    
    for i in eom.index:
        if i!='Total':
            eombr=eomfin.copy()
            eombr=eombr[eombr['REPORTING_BRANCH']==i] 
        else:
            eombr=eomfin.copy()
        m,mb=end_of(MTD,eombr,eomdisb_1,LAP='Prime+LRD')
        y,yb=end_of(ytd,eombr,eomdisb_1,LAP='Prime+LRD')
        d,db=end_of(day,eombr,eomdisb_1,LAP='Prime+LRD')
        s,sb=end_of(MTD_1,eombr,eomdisb_1,day_1,LAP='Prime+LRD')
        m_prod_br[i]=mb
        y_prod_br[i]=yb
        d_prod_br[i]=db
        stlm_prod_br[i]=sb
    
    m=[]
    d=[]
    y=[]
    s=[]
    for i in eom.index:
        if i =='Gautam Buddha Nagar':
            continue
        elif i in ['Chandigarh','Chennai','Hyderabad','Indore','Jaipur','Surat']:
            m_prod_br[i]=m_prod_br[i].drop(index='HL')
            y_prod_br[i]=y_prod_br[i].drop(index='HL')
            d_prod_br[i]=d_prod_br[i].drop(index='HL')
            stlm_prod_br[i]=stlm_prod_br[i].drop(index='HL')
            m.append(pd.concat([eom[eom.index==i],m_prod_br[i][m_prod_br[i].index!='Total']]))
            y.append(pd.concat([eoy[eoy.index==i],y_prod_br[i][y_prod_br[i].index!='Total']]))
            d.append(pd.concat([eod[eod.index==i],d_prod_br[i][d_prod_br[i].index!='Total']]))
            s.append(pd.concat([stlm[stlm.index==i],stlm_prod_br[i][stlm_prod_br[i].index!='Total']]))
        elif i == 'Total':
            m.append(eom[eom.index==i])
            y.append(eoy[eoy.index==i])
            d.append(eod[eod.index==i])
            s.append(stlm[stlm.index==i])
        else:
            m.append(pd.concat([eom[eom.index==i],m_prod_br[i][m_prod_br[i].index!='Total']]))
            y.append(pd.concat([eoy[eoy.index==i],y_prod_br[i][y_prod_br[i].index!='Total']]))
            d.append(pd.concat([eod[eod.index==i],d_prod_br[i][d_prod_br[i].index!='Total']]))
            s.append(pd.concat([stlm[stlm.index==i],stlm_prod_br[i][stlm_prod_br[i].index!='Total']]))
    mtd_df = pd.concat(m)
    ytd_df = pd.concat(y)
    ftd_df = pd.concat(d)
    stlm_df=pd.concat(s)
    
    #######################AUM#####################
    def  get_aum(base_view):
        aum=base_view[['REPORTING_BRANCH','FINTYPE','PRINCIPAL_OUTSTANDING','SUB_PRODUCT','LOAN_STATUS','LOAN_PURPOSE']]
        aum=aum[aum['LOAN_STATUS']=='Active']
        aum_hl=aum[aum['FINTYPE'].isin(['LP','NP'])==False]
        aum_lap=aum[(aum['FINTYPE'].isin(['LP','NP']))&(aum['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP'])==False)]
        aum_neolap=aum[(aum['FINTYPE'].isin(['LP','NP'])) & (aum['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP']))]
        aum_lap_b=aum_lap.groupby('REPORTING_BRANCH').sum('PRINCIPAL_OUTSTANDING')
        aum_hl_b=aum_hl.groupby('REPORTING_BRANCH').sum('PRINCIPAL_OUTSTANDING')
        aum_neolap_b=aum_neolap.groupby('REPORTING_BRANCH').sum('PRINCIPAL_OUTSTANDING')
        tot_aum=aum.groupby('REPORTING_BRANCH').sum('PRINCIPAL_OUTSTANDING')
        
        
        monbr=list(set(list(base_view.REPORTING_BRANCH.unique()))-set(aum_lap_b.index.to_list()))
        for i in monbr:
            aum_lap_b.loc[i,:]=0
        monbr=list(set(list(base_view.REPORTING_BRANCH.unique()))-set(aum_hl_b.index.to_list()))
        for i in monbr:
            aum_hl_b.loc[i,:]=0
    
        monbr=list(set(list(base_view.REPORTING_BRANCH.unique()))-set(aum_neolap_b.index.to_list()))
        for i in monbr:
            aum_neolap_b.loc[i,:]=0
        monbr=list(set(list(base_view.REPORTING_BRANCH.unique()))-set(tot_aum.index.to_list()))
        for i in monbr:
            tot_aum.loc[i,:]=0
            
        tot_aum['PRINCIPAL_OUTSTANDING']=tot_aum['PRINCIPAL_OUTSTANDING']/10000000
        aum_lap_b['PRINCIPAL_OUTSTANDING']=aum_lap_b['PRINCIPAL_OUTSTANDING']/10000000
        aum_hl_b['PRINCIPAL_OUTSTANDING']=aum_hl_b['PRINCIPAL_OUTSTANDING']/10000000
        aum_neolap_b['PRINCIPAL_OUTSTANDING']=aum_neolap_b['PRINCIPAL_OUTSTANDING']/10000000    
        aum_hl_b=aum_hl_b.sort_values(by='REPORTING_BRANCH')
        aum_hl_b.loc['In-Organic']=0
        aum_hl_b.loc['Product Total']= aum_hl_b.sum()
        aum_lap_b=aum_lap_b.sort_values(by='REPORTING_BRANCH')
        aum_lap_b.loc['In-Organic']=0
        aum_lap_b.loc['Product Total']= aum_lap_b.sum()
        aum_neolap_b=aum_neolap_b.sort_values(by='REPORTING_BRANCH')
        aum_neolap_b.loc['In-Organic']=0
        aum_neolap_b.loc['Product Total']= aum_neolap_b.sum()
        tot_aum=tot_aum.sort_values(by='REPORTING_BRANCH')
        tot_aum.loc['In-Organic']=0
        tot_aum.loc['Product Total']= tot_aum.sum()
        
        tot_aum.rename(columns={'PRINCIPAL_OUTSTANDING':'TOTAL_AUM'},inplace=True)
        aum_lap_b.rename(columns={'PRINCIPAL_OUTSTANDING':'LAP_AUM'},inplace=True)
        aum_hl_b.rename(columns={'PRINCIPAL_OUTSTANDING':'HL_AUM'},inplace=True)
        aum_neolap_b.rename(columns={'PRINCIPAL_OUTSTANDING':'NEOLAP_AUM'},inplace=True)
        df=pd.merge(aum_hl_b,aum_lap_b, on='REPORTING_BRANCH',how='inner')
        df=df.merge(aum_neolap_b, on='REPORTING_BRANCH',how='inner')
        df=df.merge(tot_aum, on='REPORTING_BRANCH',how='inner')
        df=df.round(2)
        
        return df
    aum=get_aum(base_view.copy())
    aum_targets=pd.read_excel(r"C:\Users\VAIBHAV.SRIVASTAV01\Desktop\GC_TEST\December Targets.xlsx",sheet_name='Sheet1')
    aum_targets=aum_targets[[ 'BRANCH.1', 'TOTAL AUM', 'LAP ', 'HL']]
    aum.reset_index('REPORTING_BRANCH',inplace=True)
    aum=aum.merge(aum_targets,left_on='REPORTING_BRANCH',right_on='BRANCH.1',how='left')
    aum.set_index('REPORTING_BRANCH',inplace=True)
    aum['NEOLAP TARGETS']=' '
    aum['INORG']=' '
    aum['INORGANIC_TARGETS']=' '
    
    
    aum.replace(to_replace=[np.nan,'',' '],value=0,inplace=True)
    aum=aum.round(2)
    aum[' ']=' '
    aum=aum[[' ','HL_AUM','HL', 'LAP_AUM','LAP ', 'NEOLAP_AUM', 'NEOLAP TARGETS','INORG','INORGANIC_TARGETS','TOTAL_AUM', 'TOTAL AUM']]
    
    aum.loc['In-Organic']=' '
    
    aum.reset_index('REPORTING_BRANCH',inplace=True)
    
    ################################## SUMMARY SHEET COMPLETE ###############################
    
    ################################## BRANCH DASHBOARD ###########################
    
    
    mtd_vol=eom[['LOG_VOL','SANCTION_VOL','BOOK_VOL','Disbursement_Volume','Premium']]
    ytd_vol=eoy[['LOG_VOL','SANCTION_VOL','BOOK_VOL','Disbursement_Volume','Premium']]
    ftd_vol=eod[['LOG_VOL','SANCTION_VOL','BOOK_VOL','Disbursement_Volume','Premium']]
    stlm_vol=stlm [['LOG_VOL','SANCTION_VOL','BOOK_VOL','Disbursement_Volume','Premium']]
    mtd_vol.reset_index(inplace=True)
    ftd_vol.reset_index(inplace=True)
    ytd_vol.reset_index(inplace=True)
    stlm_vol.reset_index(inplace=True)
    
    mtd_val=eom[['REQUESTED_AMOUNT','SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium', 'ROI']]
    ytd_val=eoy[['REQUESTED_AMOUNT','SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium', 'ROI']]
    ftd_val=eod[['REQUESTED_AMOUNT','SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium', 'ROI']]
    stlm_val=stlm[['REQUESTED_AMOUNT','SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium', 'ROI']]
    mtd_val['DISB']=mtd_val['DISBURSEMENT_AMOUNT_TRANCH1']+mtd_val['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    ytd_val['DISB']=ytd_val['DISBURSEMENT_AMOUNT_TRANCH1']+ytd_val['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    ftd_val['DISB']=ftd_val['DISBURSEMENT_AMOUNT_TRANCH1']+ftd_val['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    stlm_val['DISB']=stlm_val['DISBURSEMENT_AMOUNT_TRANCH1']+stlm_val['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    mtd_val=mtd_val[['REQUESTED_AMOUNT','SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium', 'ROI']]
    ytd_val=ytd_val[['REQUESTED_AMOUNT','SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium', 'ROI']]
    ftd_val=ftd_val[['REQUESTED_AMOUNT','SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium', 'ROI']]
    stlm_val=stlm_val[['REQUESTED_AMOUNT','SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium', 'ROI']]
    
    mtd_val.reset_index(inplace=True)
    ftd_val.reset_index(inplace=True)
    ytd_val.reset_index(inplace=True)
    stlm_val.reset_index(inplace=True)
    
    branch_aum=aum[['TOTAL_AUM']]
    
    adopt=eomfin.copy()
    adopt=adopt[['LAN_ID','FINTYPE','REPORTING_BRANCH','FINANCE_SOURCE_ID','SUB_CATEGORY','CUSTOMER_RESIDENTIAL_STATUS','LOGIN_STATUS','LOGIN_YEAR_MONTH']]
    adopt=adopt[adopt['LOGIN_STATUS']=='A) Login']
    hl_adopt=adopt[(adopt['FINTYPE'].isin(['HL','HT'])) & (adopt['SUB_CATEGORY'].isin(['SALARIED','NON-WORKING'])) & (adopt['CUSTOMER_RESIDENTIAL_STATUS'].isin(['NR','MN','PIO'])==False) & (adopt['LOGIN_YEAR_MONTH']==MTD[0])]
    hl_adopt['No. of Logins-API']=np.where(hl_adopt['FINANCE_SOURCE_ID']=='APIUSER',1,0)
    hl_adopt['Eligible_Logins']=1
    hl_adopt=hl_adopt[['REPORTING_BRANCH','No. of Logins-API','Eligible_Logins']]
    hl_adopt=hl_adopt.groupby('REPORTING_BRANCH').sum(['No. of Logins-API','Eligible_Logins'])
    
    monbr=list(set(list(base_view.REPORTING_BRANCH.unique()))-set(hl_adopt.index.to_list()))
    for i in monbr:
        hl_adopt.loc[i,:]=0
    hl_adopt=hl_adopt.sort_values('REPORTING_BRANCH')
    hl_adopt.loc['Total']=hl_adopt.sum()
    hl_adopt['Adoption rate (%) *']=hl_adopt['No. of Logins-API']/hl_adopt['Eligible_Logins']
    hl_adopt.replace(to_replace=[np.nan,'',' '],value=0,inplace=True)
    
    
    
    lap_adopt=adopt[(adopt['FINTYPE'].isin(['LP','NP']))& (adopt['CUSTOMER_RESIDENTIAL_STATUS'].isin(['NR','MN','PIO'])==False) & (adopt['LOGIN_YEAR_MONTH']==MTD[0])]
    lap_adopt['No. of Logins-API']=np.where(lap_adopt['FINANCE_SOURCE_ID']=='APIUSER',1,0)
    lap_adopt['Eligible_Logins']=1
    lap_adopt=lap_adopt[['REPORTING_BRANCH','No. of Logins-API','Eligible_Logins']]
    lap_adopt=lap_adopt.groupby('REPORTING_BRANCH').sum(['No. of Logins-API','Eligible_Logins'])
    monbr=list(set(list(base_view.REPORTING_BRANCH.unique()))-set(lap_adopt.index.to_list()))
    for i in monbr:
        lap_adopt.loc[i,:]=0
    lap_adopt=lap_adopt.sort_values('REPORTING_BRANCH')
    lap_adopt.loc['Total']=lap_adopt.sum()
    lap_adopt['Adoption rate (%) *']=lap_adopt['No. of Logins-API']/lap_adopt['Eligible_Logins']
    lap_adopt.replace(to_replace=[np.nan,'',' '],value=0,inplace=True)
    
    adoption=pd.merge(hl_adopt,lap_adopt,on='REPORTING_BRANCH',how='left')
    adoption.reset_index(inplace=True)
    ############################### BRANCH DASHBOARD COMPLETE #################### 
    ################################## PRODUCT DASHBOARD ###########################
    
    
    # mtd_volp=eomp[['LOG_VOL','SANCTION_VOL','BOOK_VOL','Disbursement_Volume','Pmium']]
    # ytd_volp=eoyp[['LOG_VOL','SANCTION_VOL','BOOK_VOL','Disbursement_Volume','Premium']]
    # ftd_volp=eodp[['LOG_VOL','SANCTION_VOL','BOOK_VOL','Disbursement_Volume','Premium']]
    # stlm_volp=stlmp[['LOG_VOL','SANCTION_VOL','BOOK_VOL','Disbursement_Volume','Premium']]
    
    # mtd_volp.reset_index(inplace=True)
    # ftd_volp.reset_index(inplace=True)
    # ytd_volp.reset_index(inplace=True)
    # stlm_volp.reset_index(inplace=True)
    
    # mtd_valp=eomp[['REQUESTED_AMOUNT','SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium', 'ROI']]
    # ytd_valp=eoyp[['REQUESTED_AMOUNT','SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium', 'ROI']]
    # ftd_valp=eodp[['REQUESTED_AMOUNT','SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium', 'ROI']]
    # stlm_valp=stlmp[['REQUESTED_AMOUNT','SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium', 'ROI']]
    # mtd_valp['DISB']=mtd_valp['DISBURSEMENT_AMOUNT_TRANCH1']+mtd_valp['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    # ytd_valp['DISB']=ytd_valp['DISBURSEMENT_AMOUNT_TRANCH1']+ytd_valp['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    # ftd_valp['DISB']=ftd_valp['DISBURSEMENT_AMOUNT_TRANCH1']+ftd_valp['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    # stlm_valp['DISB']=stlm_valp['DISBURSEMENT_AMOUNT_TRANCH1']+stlm_valp['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    # mtd_valp=mtd_valp[['REQUESTED_AMOUNT','SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium', 'ROI']]
    # ytd_valp=ytd_valp[['REQUESTED_AMOUNT','SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium', 'ROI']]
    # ftd_valp=ftd_valp[['REQUESTED_AMOUNT','SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium', 'ROI']]
    # stlm_valp=stlm_valp[['REQUESTED_AMOUNT','SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium', 'ROI']]
    # mtd_valp.reset_index(inplace=True)
    # ftd_valp.reset_index(inplace=True)
    # ytd_valp.reset_index(inplace=True)
    # stlm_valp.reset_index(inplace=True)
    def prod(MTD,eomfin,eomdisb_1,lastmonth=None,LRD=None):
        eomfin_book=eomfin.copy()
        if type(MTD)==datetime.datetime:
            eomfin_book['bookdate']=np.where(True,pd.to_datetime(eomfin_book['BOOKING_DATE']).dt.date ,0)
            eomfin_book=eomfin_book[eomfin_book['bookdate']==(MTD.date())]
            # eomfin_book['BOOK_VOL']= np.where(eomfin_book['bookdate']==(MTD.date()),1,0)
        elif lastmonth!=None:
            eomfin_book['bookdate']=np.where(True,pd.to_datetime(eomfin_book['BOOKING_DATE']).dt.date ,0)
            eomfin_book=eomfin_book[(eomfin_book['bookdate']<=(lastmonth.date())) & (eomfin_book['BOOK_YEAR_MONTH'].isin(MTD))]
        else:
            eomfin_book=eomfin_book[eomfin_book['BOOK_YEAR_MONTH'].isin(MTD)]
        
            # eomfin_book['BOOK_VOL']= np.where((eomfin_book['BOOK_YEAR_MONTH'].isin(MTD)),1,0)
        eomfin_book=eomfin_book[eomfin_book['STATUS']=='Booked']
        eomfin_book['BOOK_VOL']= np.where(eomfin_book['STATUS']=='Booked',1,0)
        eomfin_book['INS_VOL']=np.where(eomfin_book['INS_FLAG']=='Y',1,0)
        eomfin_book['BOOKING_AMOUNT']=eomfin_book['BOOKING_AMOUNT'].round(2)
        eomfin_book['WROI']=eomfin_book['ROI']*eomfin_book['BOOKING_AMOUNT']
        eomfin_book=eomfin_book[[ 'LAN_ID','REPORTING_BRANCH', 'BOOKING_AMOUNT','BOOK_VOL','FINTYPE','WROI','NET_PREMIUM','GPLFLAG_SANCTIONS','SUB_PRODUCT','INS_VOL']]
        eomfin_bookview=eomfin_book.groupby('REPORTING_BRANCH').sum(['BOOKING_AMOUNT','BOOK_VOL','WROI','INS_FLAG'])
        eomfin_bookview['ROI']=eomfin_bookview['WROI']/eomfin_bookview['BOOKING_AMOUNT']
        eomfin_bookview['Premium']=eomfin_bookview['NET_PREMIUM']/10000000
        eomfin_bookview['Penetration']=eomfin_bookview['NET_PREMIUM']/eomfin_bookview['BOOKING_AMOUNT']
        eomfin_bookview['BOOKING_AMOUNT']=eomfin_bookview['BOOKING_AMOUNT']/10000000
        eomfin_bookview=eomfin_bookview[[  'BOOKING_AMOUNT','BOOK_VOL','ROI','Premium','Penetration']]
        eomfin_book['Prod']=np.where(((eomfin_book['FINTYPE'].isin(['LP'])) & ((eomfin_book['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP',LRD])==False))),'Prime LAP',np.where(((eomfin_book['FINTYPE'].isin(['NP'])) & ((eomfin_book['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP',LRD])==False))),'NRP',np.where((eomfin_book['FINTYPE'].isin(['LP','NP']) & (eomfin_book['SUB_PRODUCT'].isin([LRD]))),LRD,np.where((eomfin_book['FINTYPE'].isin(['LP','NP']) & (eomfin_book['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP']))),'NEO-LAP',np.where((eomfin_book['FINTYPE'].isin(['LP','NP'])==False) & (eomfin_book['FINTYPE'].isin(['FL','FT'])),'FLEXI',np.where((eomfin_book['FINTYPE'].isin(['LP','NP','FL','FT'])==False) & (eomfin_book['GPLFLAG_SANCTIONS']=='GPL'),'GPL','NON GPL' ))))))
        eomfin_pbview=eomfin_book.groupby('Prod').sum(['BOOKING_AMOUNT','BOOK_VOL','WROI','INS_VOL'])
        eomfin_pbview['ROI']=eomfin_pbview['WROI']/eomfin_pbview['BOOKING_AMOUNT']
        eomfin_pbview['Premium']=eomfin_pbview['NET_PREMIUM']/10000000
        eomfin_pbview['Penetration']=eomfin_pbview['NET_PREMIUM']/eomfin_pbview['BOOKING_AMOUNT']
        eomfin_pbview['BOOKING_AMOUNT']=eomfin_pbview['BOOKING_AMOUNT']/10000000
        eomfin_pbview=eomfin_pbview[[  'BOOKING_AMOUNT','BOOK_VOL','ROI','Premium','Penetration','INS_VOL']]
        
        eomfin_log=eomfin.copy()
        
        if type(MTD)==datetime.datetime:
            eomfin_log['logdate']=np.where(True,pd.to_datetime(eomfin_log['EOMLOGN']).dt.date ,0)
            eomfin_log=eomfin_log[eomfin_log['logdate']==(MTD.date())]
            # eomfin_log['LOG_VOL']= np.where(eomfin_log['logdate']==(MTD.date()),1,0)
        elif lastmonth!=None:
            eomfin_log['logdate']=np.where(True,pd.to_datetime(eomfin_log['EOMLOGN']).dt.date ,0)
            eomfin_log=eomfin_log[(eomfin_log['logdate']<=(lastmonth.date())) & (eomfin_log['LOGIN_YEAR_MONTH'].isin(MTD))]
        else:
            eomfin_log=eomfin_log[eomfin_log['LOGIN_YEAR_MONTH'].isin(MTD)]
            # eomfin_log['LOG_VOL']= np.where((eomfin_log['LOGIN_YEAR_MONTH'].isin(MTD)),1,0)
        eomfin_log['LOG_VOL']= np.where(eomfin_log['LOGIN_STATUS']=='A) Login',1,0)
        eomfin_log=eomfin_log[eomfin_log['LOGIN_STATUS']=='A) Login']
        eomfin_log['REQUESTED_AMOUNT']=eomfin_log['REQUESTED_AMOUNT']/10000000
        eomfin_log['REQUESTED_AMOUNT']=eomfin_log['REQUESTED_AMOUNT'].round(2)
        eomfin_log=eomfin_log[[ 'LAN_ID', 'REPORTING_BRANCH','REQUESTED_AMOUNT','LOG_VOL','FINTYPE','GPLFLAG_SANCTIONS','SUB_PRODUCT']]
        eomfin_logview=eomfin_log.groupby('REPORTING_BRANCH').sum(['REQUESTED_AMOUNT','LOG_VOL'])
        eomfin_log['Prod']=np.where(((eomfin_log['FINTYPE'].isin(['LP'])) & ((eomfin_log['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP',LRD])==False))),'Prime LAP',np.where(((eomfin_log['FINTYPE'].isin(['NP'])) & ((eomfin_log['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP',LRD])==False))),'NRP',np.where((eomfin_log['FINTYPE'].isin(['LP','NP']) & (eomfin_log['SUB_PRODUCT'].isin([LRD]))),LRD,np.where((eomfin_log['FINTYPE'].isin(['LP','NP']) & (eomfin_log['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP']))),'NEO-LAP',np.where((eomfin_log['FINTYPE'].isin(['LP','NP'])==False) & (eomfin_log['FINTYPE'].isin(['FL','FT'])),'FLEXI',np.where((eomfin_log['FINTYPE'].isin(['LP','NP','FL','FT'])==False) & (eomfin_log['GPLFLAG_SANCTIONS']=='GPL'),'GPL','NON GPL' ))))))
        eomfin_plview=eomfin_log.groupby('Prod').sum(['REQUESTED_AMOUNT','LOG_VOL'])
        
        
        eomfin_sanc=eomfin.copy()
        
        if type(MTD)==datetime.datetime:
            eomfin_sanc['sancdate']=np.where(True,pd.to_datetime(eomfin_sanc['EOMSNCTN']).dt.date ,0)
            eomfin_sanc=eomfin_sanc[eomfin_sanc['sancdate']==(MTD.date())]
            # eomfin_sanc['SANCTION_VOL']= np.where((eomfin_sanc['sancdate']==(MTD.date())),1,0)
        elif lastmonth!=None:
            eomfin_sanc['sancdate']=np.where(True,pd.to_datetime(eomfin_sanc['EOMSNCTN']).dt.date ,0)
            eomfin_sanc=eomfin_sanc[(eomfin_sanc['sancdate']<=(lastmonth.date())) & (eomfin_sanc['FS_YEAR_MONTH'].isin(MTD))]
        else:
            eomfin_sanc=eomfin_sanc[eomfin_sanc['FS_YEAR_MONTH'].isin(MTD)]
            # eomfin_sanc['SANCTION_VOL']= np.where((eomfin_sanc['FS_YEAR_MONTH'].isin(MTD)),1,0)
        eomfin_sanc['SANCTION_VOL']= np.where(eomfin_sanc['STATUS_SEG']=='A) Final Sanction',1,0)
        eomfin_sanc=eomfin_sanc[eomfin_sanc['STATUS_SEG']=='A) Final Sanction']
        eomfin_sanc['SANCTION_AMOUNT']=eomfin_sanc['SANCTION_AMOUNT']/10000000
        eomfin_sanc['SANCTION_AMOUNT']=eomfin_sanc['SANCTION_AMOUNT'].round(2)
        eomfin_sanc=eomfin_sanc[[ 'LAN_ID','REPORTING_BRANCH', 'SANCTION_AMOUNT','SANCTION_VOL','FINTYPE','GPLFLAG_SANCTIONS','SUB_PRODUCT']]
        eomfin_sancview=eomfin_sanc.groupby('REPORTING_BRANCH').sum(['SANCTION_AMOUNT','SANCTION_VOL'])
    
        eomfin_sanc['Prod']=np.where(((eomfin_sanc['FINTYPE'].isin(['LP'])) & ((eomfin_sanc['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP',LRD])==False))),'Prime LAP',np.where(((eomfin_sanc['FINTYPE'].isin(['NP'])) & ((eomfin_sanc['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP',LRD])==False))),'NRP',np.where((eomfin_sanc['FINTYPE'].isin(['LP','NP']) & (eomfin_sanc['SUB_PRODUCT'].isin([LRD]))),LRD,np.where((eomfin_sanc['FINTYPE'].isin(['LP','NP']) & (eomfin_sanc['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP']))),'NEO-LAP',np.where((eomfin_sanc['FINTYPE'].isin(['LP','NP'])==False) & (eomfin_sanc['FINTYPE'].isin(['FL','FT'])),'FLEXI',np.where((eomfin_sanc['FINTYPE'].isin(['LP','NP','FL','FT'])==False) & (eomfin_sanc['GPLFLAG_SANCTIONS']=='GPL'),'GPL','NON GPL' ))))))
        eomfin_psview=eomfin_sanc.groupby('Prod').sum(['SANCTION_AMOUNT','SANCTION_VOL'])
        
        eomfin_insanc=eomfin.copy()
        
        if type(MTD)==datetime.datetime:
            eomfin_insanc['inc_sancdate']=np.where(True,pd.to_datetime(eomfin_insanc['EOMCLTRL']).dt.date ,0)
            eomfin_insanc=eomfin_insanc[eomfin_insanc['inc_sancdate']==(MTD.date())]
            # eomfin_insanc['SANCTION_VOL']= np.where((eomfin_insanc['sancdate']==(MTD.date())),1,0)
        elif lastmonth!=None:
            eomfin_insanc['inc_sancdate']=np.where(True,pd.to_datetime(eomfin_insanc['EOMCLTRL']).dt.date ,0)
            eomfin_insanc=eomfin_insanc[(eomfin_insanc['inc_sancdate']<=(lastmonth.date())) & (eomfin_insanc['IS_YEAR_MONTH'].isin(MTD))]
        else:
            eomfin_insanc=eomfin_insanc[eomfin_insanc['IS_YEAR_MONTH'].isin(MTD)]
            # eomfin_insanc['SANCTION_VOL']= np.where((eomfin_insanc['FS_YEAR_MONTH'].isin(MTD)),1,0)
        eomfin_insanc['IN_SANCTION_VOL']= np.where(eomfin_insanc['STATUS_SEG']=='C) Income Sanction',1,0)
        eomfin_insanc=eomfin_insanc[eomfin_insanc['STATUS_SEG']=='C) Income Sanction']
        eomfin_insanc['IN_SANCTION_AMOUNT']=eomfin_insanc['SANCTION_AMOUNT']/10000000
        eomfin_insanc['IN_SANCTION_AMOUNT']=eomfin_insanc['IN_SANCTION_AMOUNT'].round(2)
        eomfin_insanc=eomfin_insanc[[ 'LAN_ID','REPORTING_BRANCH', 'IN_SANCTION_AMOUNT','IN_SANCTION_VOL','FINTYPE','GPLFLAG_SANCTIONS','SUB_PRODUCT']]
        eomfin_insancview=eomfin_insanc.groupby('REPORTING_BRANCH').sum(['IN_SANCTION_AMOUNT','IN_SANCTION_VOL'])
        eomfin_insanc['Prod']=np.where(((eomfin_insanc['FINTYPE'].isin(['LP'])) & ((eomfin_insanc['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP',LRD])==False))),'Prime LAP',np.where(((eomfin_insanc['FINTYPE'].isin(['NP'])) & ((eomfin_insanc['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP',LRD])==False))),'NRP',np.where((eomfin_insanc['FINTYPE'].isin(['LP','NP']) & (eomfin_insanc['SUB_PRODUCT'].isin([LRD]))),LRD,np.where((eomfin_insanc['FINTYPE'].isin(['LP','NP']) & (eomfin_insanc['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP']))),'NEO-LAP',np.where((eomfin_insanc['FINTYPE'].isin(['LP','NP'])==False) & (eomfin_insanc['FINTYPE'].isin(['FL','FT'])),'FLEXI',np.where((eomfin_insanc['FINTYPE'].isin(['LP','NP','FL','FT'])==False) & (eomfin_insanc['GPLFLAG_SANCTIONS']=='GPL'),'GPL','NON GPL' ))))))
        eomfin_inpsview=eomfin_insanc.groupby('Prod').sum(['IN_SANCTION_AMOUNT','IN_SANCTION_VOL'])
        
        eomfin_reject=eomfin.copy()
        
        if type(MTD)==datetime.datetime:
            eomfin_reject['rejectdate']=np.where(True,pd.to_datetime(eomfin_reject['EOMRJCT']).dt.date ,0)
            eomfin_reject=eomfin_reject[eomfin_reject['rejectdate']==(MTD.date())]
            # eomfin_reject['SANCTION_VOL']= np.where((eomfin_reject['sancdate']==(MTD.date())),1,0)
        elif lastmonth!=None:
            eomfin_reject['rejectdate']=np.where(True,pd.to_datetime(eomfin_reject['EOMRJCT']).dt.date ,0)
            eomfin_reject=eomfin_reject[(eomfin_reject['rejectdate']<=(lastmonth.date())) & (eomfin_reject['REJECT_YEAR_MONTH'].isin(MTD))]
        else:
            eomfin_reject=eomfin_reject[eomfin_reject['REJECT_YEAR_MONTH'].isin(MTD)]
            # eomfin_reject['SANCTION_VOL']= np.where((eomfin_reject['FS_YEAR_MONTH'].isin(MTD)),1,0)
        eomfin_reject['REJECT_VOL']= np.where(eomfin_reject['STATUS_SEG']=='B) Rejected',1,0)
        eomfin_reject=eomfin_reject[eomfin_reject['STATUS_SEG']=='B) Rejected']
        eomfin_reject['REJECT_AMOUNT']=eomfin_reject['REQUESTED_AMOUNT']/10000000
        eomfin_reject['REJECT_AMOUNT']=eomfin_reject['REJECT_AMOUNT'].round(2)
        eomfin_reject=eomfin_reject[[ 'LAN_ID','REPORTING_BRANCH', 'REJECT_AMOUNT','REJECT_VOL','FINTYPE','GPLFLAG_SANCTIONS','SUB_PRODUCT']]
        eomfin_rejectview=eomfin_reject.groupby('REPORTING_BRANCH').sum(['REJECT_AMOUNT','REJECT_VOL'])
        eomfin_reject['Prod']=np.where(((eomfin_reject['FINTYPE'].isin(['LP'])) & ((eomfin_reject['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP',LRD])==False))),'Prime LAP',np.where(((eomfin_reject['FINTYPE'].isin(['NP'])) & ((eomfin_reject['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP',LRD])==False))),'NRP',np.where((eomfin_reject['FINTYPE'].isin(['LP','NP']) & (eomfin_reject['SUB_PRODUCT'].isin([LRD]))),LRD,np.where((eomfin_reject['FINTYPE'].isin(['LP','NP']) & (eomfin_reject['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP']))),'NEO-LAP',np.where((eomfin_reject['FINTYPE'].isin(['LP','NP'])==False) & (eomfin_reject['FINTYPE'].isin(['FL','FT'])),'FLEXI',np.where((eomfin_reject['FINTYPE'].isin(['LP','NP','FL','FT'])==False) & (eomfin_reject['GPLFLAG_SANCTIONS']=='GPL'),'GPL','NON GPL' ))))))
        
        eomfin_rejectpsview=eomfin_reject.groupby('Prod').sum(['REJECT_AMOUNT','REJECT_VOL'])
        
        
        
        m1=pd.merge(eomfin_bookview,eomfin_logview,left_on=('REPORTING_BRANCH'),right_on=('REPORTING_BRANCH'), how='outer')
        m2=pd.merge(m1,eomfin_sancview,left_on=('REPORTING_BRANCH'),right_on=('REPORTING_BRANCH'), how='outer')
        m3=pd.merge(m2,eomfin_insancview,left_on=('REPORTING_BRANCH'),right_on=('REPORTING_BRANCH'), how='outer')
        eomfinview=pd.merge(m3,eomfin_rejectview,left_on=('REPORTING_BRANCH'),right_on=('REPORTING_BRANCH'), how='outer')
        
        
        t1=pd.merge(eomfin_pbview,eomfin_plview,left_on=('Prod'),right_on=('Prod'), how='outer')
        t2=pd.merge(t1,eomfin_psview,left_on=('Prod'),right_on=('Prod'), how='outer')
        t3=pd.merge(t2,eomfin_inpsview,left_on=('Prod'),right_on=('Prod'), how='outer')
        eompfinview=pd.merge(t3,eomfin_rejectpsview,left_on=('Prod'),right_on=('Prod'), how='outer')
        
        eomdisb=pd.merge(eomfin,eomdisb_1,on=('LAN_ID'),how='left')
        # eomdisb=eomdisb_1.copy()
        if type(MTD)==datetime.datetime:
            eomdisb['DISBDATE']=np.where(True,pd.to_datetime(eomdisb['DISBURSEMENT_DATE']).dt.date ,0)
            eomdisb=eomdisb[eomdisb['DISBDATE']==(MTD.date())]
            # eomdisb['DISB_VOL']=np.where((eomdisb['DISBDATE']==(MTD.date())),1,0)
        elif lastmonth!=None:
            eomdisb['DISBDATE']=np.where(True,pd.to_datetime(eomdisb['DISBURSEMENT_DATE']).dt.date ,0)
            eomdisb=eomdisb[(eomdisb['DISBDATE']<=(lastmonth.date())) & (eomdisb['DISB_YEAR_MONTH'].isin(MTD))]
        else:
            eomdisb=eomdisb[eomdisb['DISB_YEAR_MONTH'].isin(MTD)]
            # eomdisb['DISB_VOL']=np.where((eomdisb['DISB_YEAR_MONTH'].isin(MTD)),1,0)
        eomdisb['DISB_VOL']=np.where(True,1,0)
        eomdisb['DISBURSEMENT_AMOUNT_TRANCH1']=np.where(eomdisb['DISBURSEMENT_SEQUENCE']==1,eomdisb['DISBURSEMENT_AMOUNT']/1000000000,0)
        eomdisb['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']=np.where(eomdisb['DISBURSEMENT_SEQUENCE']!=1,eomdisb['DISBURSEMENT_AMOUNT']/1000000000,0)
        eomdisb['DISBURSEMENT_AMOUNT_TRANCH1']=eomdisb['DISBURSEMENT_AMOUNT_TRANCH1'].round(2)
        eomdisb['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']=eomdisb['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES'].round(2)
        eomdisb=eomdisb[[ 'REPORTING_BRANCH', 'DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES','FINTYPE','GPLFLAG_SANCTIONS','SUB_PRODUCT']]
        eomdisbview=eomdisb.groupby('REPORTING_BRANCH').sum(['DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES'])
        eomdisb['Prod']=np.where(((eomdisb['FINTYPE'].isin(['LP'])) & ((eomdisb['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP',LRD])==False))),'Prime LAP',np.where(((eomdisb['FINTYPE'].isin(['NP'])) & ((eomdisb['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP',LRD])==False))),'NRP',np.where((eomdisb['FINTYPE'].isin(['LP','NP']) & (eomdisb['SUB_PRODUCT'].isin([LRD]))),LRD,np.where((eomdisb['FINTYPE'].isin(['LP','NP']) & (eomdisb['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP']))),'NEO-LAP',np.where((eomdisb['FINTYPE'].isin(['LP','NP'])==False) & (eomdisb['FINTYPE'].isin(['FL','FT'])),'FLEXI',np.where((eomdisb['FINTYPE'].isin(['LP','NP','FL','FT'])==False) & (eomdisb['GPLFLAG_SANCTIONS']=='GPL'),'GPL','NON GPL' ))))))
        eomfin_pdview=eomdisb.groupby('Prod').sum(['DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES'])
        
        eomp=pd.merge(eompfinview,eomfin_pdview,left_on=('Prod'),right_on=('Prod'), how='outer')
        eom=pd.merge(eomfinview,eomdisbview,left_on=('REPORTING_BRANCH'),right_on=('REPORTING_BRANCH'), how='outer')
        eom=eom.sort_values(by='REPORTING_BRANCH')
        eomp=eomp.sort_values(by='Prod')
        eom=eom[['LOG_VOL','REQUESTED_AMOUNT','SANCTION_VOL','SANCTION_AMOUNT','BOOK_VOL','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES','REJECT_AMOUNT','REJECT_VOL','IN_SANCTION_AMOUNT','IN_SANCTION_VOL','ROI','Premium','Penetration']]
        eomp=eomp[['LOG_VOL','REQUESTED_AMOUNT','SANCTION_VOL','SANCTION_AMOUNT','BOOK_VOL','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES','REJECT_AMOUNT','REJECT_VOL','IN_SANCTION_AMOUNT','IN_SANCTION_VOL','ROI','Premium','Penetration','INS_VOL']]
        
        monbr=list(set(list(base_view.REPORTING_BRANCH.unique()))-set(eom.index.to_list()))
        if LRD==None:
            missprod=list(set(['HL','GPL','NON GPL','FLEXI', 'LAP','NEO-LAP','NRP','Prime LAP'])-set(eomp.index.to_list()))
        else:
            missprod=list(set(['HL','GPL','NON GPL','FLEXI', 'LAP','NEO-LAP','NRP','Prime LAP',LRD])-set(eomp.index.to_list()))
        for i in monbr:
            eom.loc[i,:]=0
        for i in missprod:
            eomp.loc[i,:]=0
        eom=eom.sort_values(by='REPORTING_BRANCH')
        
        eomp=eomp.reindex(['HL','GPL','NON GPL','FLEXI', 'LAP','NEO-LAP','NRP','Prime LAP',LRD])
        eom.replace(to_replace=[np.nan,'',' '],value=0,inplace=True)
        eomp.replace(to_replace=[np.nan,'',' '],value=0,inplace=True)
        a=eom['ROI']*eom['BOOKING_AMOUNT']
        a=pd.DataFrame(a)
        a['BOOKING_AMOUNT']=eom['BOOKING_AMOUNT']
        a['NET_PREMIUM']=eom['Premium']
        a.loc['Total']= a.sum()
        eom.loc['Total']= eom.sum()
        eom.loc['Total','ROI']=a.loc['Total',0]/a.loc['Total','BOOKING_AMOUNT']
        eom.loc['Total','Penetration']=a.loc['Total','NET_PREMIUM']/a.loc['Total','BOOKING_AMOUNT']
        eomp.loc['Total']= eomp.sum()
        eomp.loc['Total','ROI']=a.loc['Total',0]/a.loc['Total','BOOKING_AMOUNT']
        eomp.loc['Total','Penetration']=a.loc['Total','NET_PREMIUM']/a.loc['Total','BOOKING_AMOUNT']
        eom['Disbursement_Volume']=eom['BOOK_VOL']
        eomp['Disbursement_Volume']=eomp['BOOK_VOL']
        
        eom=eom[['LOG_VOL','REQUESTED_AMOUNT','SANCTION_VOL','SANCTION_AMOUNT','BOOK_VOL','BOOKING_AMOUNT','Disbursement_Volume','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES','REJECT_AMOUNT','REJECT_VOL','IN_SANCTION_AMOUNT','IN_SANCTION_VOL','ROI','Premium','Penetration']]
        eomp=eomp[['LOG_VOL','REQUESTED_AMOUNT','SANCTION_VOL','SANCTION_AMOUNT','BOOK_VOL','BOOKING_AMOUNT','Disbursement_Volume','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES','REJECT_AMOUNT','REJECT_VOL','IN_SANCTION_AMOUNT','IN_SANCTION_VOL','ROI','Premium','Penetration','INS_VOL']]
        eom.replace(to_replace=[np.nan,'',' '],value=0,inplace=True)
        eomp.replace(to_replace=[np.nan,'',' '],value=0,inplace=True)
        eomp.loc['HL']=eomp.loc['GPL']+eomp.loc['NON GPL']
        eomp.loc['LAP']=eomp.loc['NEO-LAP']+eomp.loc['NRP']+eomp.loc['Prime LAP']+eomp.loc['LRD']
        eomp.loc['HL','ROI']=((eomp.loc['GPL','ROI']*eomp.loc['GPL','BOOKING_AMOUNT']) + (eomp.loc['NON GPL','ROI']*eomp.loc['NON GPL','BOOKING_AMOUNT']))/eomp.loc['HL','BOOKING_AMOUNT']
        eomp.loc['LAP','ROI']=((eomp.loc['NEO-LAP','ROI']*eomp.loc['NEO-LAP','BOOKING_AMOUNT']) + (eomp.loc['LRD','ROI']*eomp.loc['LRD','BOOKING_AMOUNT'])+ (eomp.loc['NRP','ROI']*eomp.loc['NRP','BOOKING_AMOUNT'])+ (eomp.loc['Prime LAP','ROI']*eomp.loc['Prime LAP','BOOKING_AMOUNT']))/eomp.loc['LAP','BOOKING_AMOUNT']
        eomp['DISB']=eomp['DISBURSEMENT_AMOUNT_TRANCH1']+eomp['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
        vol=eomp[['LOG_VOL', 'SANCTION_VOL','BOOK_VOL','BOOK_VOL','INS_VOL']]
        val=eomp[['REQUESTED_AMOUNT','SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium', 'ROI']]
        val.reset_index(inplace=True)
        vol.reset_index(inplace=True)
        return vol,val
    
    pmvol,pmval=prod(MTD,eomfin,eomdisb_1,LRD='LRD')
    pyvol,pyval=prod(ytd,eomfin,eomdisb_1,LRD='LRD')
    pdvol,pdval=prod(day,eomfin,eomdisb_1,LRD='LRD')
    psvol,psval=prod(MTD_1,eomfin,eomdisb_1,day_1,LRD='LRD')
    ############################### PRODUCT DASHBOARD COMPLETE ####################
    
    
    ############################### INCOME SANCTION FLOW #########################
    sanc=eomfin.copy()
    isflow=sanc[[ 'REPORTING_BRANCH', 'SANCTION_AMOUNT','FINTYPE','IS_YEAR_MONTH','STATUS_SEG']]
    isflow=isflow[isflow['STATUS_SEG']=='C) Income Sanction']
    isflow['IN_SANCTION_VOL']= np.where(isflow['STATUS_SEG']=='C) Income Sanction',1,0)
    ltd_sanc=isflow.copy()
    isflow=isflow[isflow['IS_YEAR_MONTH'].isin(['1900-1','2020-11','2021-1', '2021-2','2021-3','2020-12'])==False]
    
    def sanctionflow(flows,MTD):
        isflow=flows.copy()
        isflow=isflow[isflow['FINTYPE']=='HL']
        flow={}
        for i in isflow['IS_YEAR_MONTH'].unique():
            flowis=isflow.copy()
            flowis=flowis[flowis['IS_YEAR_MONTH']==i]
            flowis=flowis[['REPORTING_BRANCH','SANCTION_AMOUNT','IN_SANCTION_VOL']]
            flowis=flowis.groupby('REPORTING_BRANCH').sum('SANCTION_AMOUNT','IN_SANCTION_VOL')
            monbr=list(set(list(base_view.REPORTING_BRANCH.unique()))-set(flowis.index.to_list()))
            for j in monbr:
                flowis.loc[j,:]=0
            flowis=flowis.sort_values(by='REPORTING_BRANCH')
            flowis.loc['Total']=flowis.sum()
            flow[i]=flowis
        a=['REPORTING_BRANCH','2021-4','2021-5','2021-6','2021-7','2021-8','2021-9','2021-10','2021-11','2021-12','2022-1','2022-2','2022-3','2022-4','2022-5','2022-6','2022-7','2022-8','2022-9','2022-10','2022-11']
        a=a+ list(set(list(flow.keys()))-set(a))
        sancflow_val=pd.DataFrame(columns=a)
        sancflow_vol=pd.DataFrame(columns=a)
        for x in list(base_view.REPORTING_BRANCH.unique()):
            # print(x)
            sancflow_val.loc[x,'REPORTING_BRANCH']=x
            sancflow_vol.loc[x,'REPORTING_BRANCH']=x
        sancflow_val.set_index('REPORTING_BRANCH',inplace=True)
        sancflow_vol.set_index('REPORTING_BRANCH',inplace=True)
        for i in sancflow_val.columns:
            for j in base_view.REPORTING_BRANCH.unique():
                sancflow_val.loc[j,i]=flow[i].loc[j,'SANCTION_AMOUNT']
                sancflow_vol.loc[j,i]=flow[i].loc[j,'IN_SANCTION_VOL']
        sancflow_val=sancflow_val.sort_values(by='REPORTING_BRANCH')  
        sancflow_vol=sancflow_vol.sort_values(by='REPORTING_BRANCH') 
        sancflow_val.loc['Total']=sancflow_val.sum()
        sancflow_vol.loc['Total']=sancflow_vol.sum()
        sancflow_val=sancflow_val/10000000
        
        return sancflow_val,sancflow_vol
    
    sanc_val,sanc_vol=sanctionflow(isflow,MTD)
    
    ltd_sanc=ltd_sanc.groupby('REPORTING_BRANCH').sum('SANCTION_AMOUNT','IN_SANCTION_VOL')
    ltd_sanc=ltd_sanc[['SANCTION_AMOUNT','IN_SANCTION_VOL']]
    ltd_sanc=ltd_sanc.sort_values(by='REPORTING_BRANCH') 
    ltd_sanc.loc['Total']=ltd_sanc.sum() 
    ltd_sanc['SANCTION_AMOUNT']=ltd_sanc['SANCTION_AMOUNT']/10000000
    val=ltd_sanc[['SANCTION_AMOUNT']]
    vol=ltd_sanc[['IN_SANCTION_VOL']]
    sanc_val=sanc_val.merge(val,on='REPORTING_BRANCH',how='left')
    sanc_vol=sanc_vol.merge(vol,on='REPORTING_BRANCH',how='left')
    sanc_val.replace(to_replace=[np.nan,'',' '],value=0,inplace=True)
    sanc_vol.replace(to_replace=[np.nan,'',' '],value=0,inplace=True)
    sanc_val=sanc_val.astype(int)
    sanc_vol=sanc_vol.astype(int)
    
    ########################### SANCTION FLOW COMPLETED ##########################
    ########################### AUM PRODUCT TREND ###############################
    
    def  get_aum_prod(base_view):
        aum_tdf=base_view[['REPORTING_BRANCH','FINTYPE','PRINCIPAL_OUTSTANDING','LOAN_STATUS']]
        aum_tdf=aum_tdf[aum_tdf['LOAN_STATUS']=='Active']
        aum_tdf['PRINCIPAL_OUTSTANDING']=aum_tdf['PRINCIPAL_OUTSTANDING']/10000000
        aum_prod={}
        aum_prod['Total']=aum_tdf.groupby('REPORTING_BRANCH').sum('PRINCIPAL_OUTSTANDING')
        aum_prod['Total'].loc['Total']= aum_prod['Total'].sum()
        for i in aum_tdf.FINTYPE.unique():
            aum=aum_tdf.copy()
            aum=aum[aum['FINTYPE'].isin([i])]
            aum_df=aum.groupby('REPORTING_BRANCH').sum('PRINCIPAL_OUTSTANDING')
            monbr=list(set(list(base_view.REPORTING_BRANCH.unique()))-set(aum_df.index.to_list()))
            for z in monbr:
                aum_df.loc[z,:]=0
            aum_df=aum_df.sort_values(by='REPORTING_BRANCH')
            
            aum_df.loc['Total']= aum_df.sum()
            aum_df.rename(columns={'PRINCIPAL_OUTSTANDING':i},inplace=True)
            aum_prod[i]=aum_df
            
        dfs=aum_prod.values()
        merged_df = reduce(lambda l, r: pd.merge(l, r, on='REPORTING_BRANCH', how='inner'), dfs)
        merged_df=merged_df.round(2)
        return merged_df
    aum_product_trend=get_aum_prod(base_view.copy())
    aum_product_trend=aum_product_trend[['FL', 'HL', 'HT', 'LP', 'NP', 'LT', 'FT','PRINCIPAL_OUTSTANDING']]
    aum_product_trend=aum_product_trend.transpose()
    aum_product_trend['In-Organic']=' '
    aum_product_trend=aum_product_trend[['Ahmedabad', 'Bangalore', 'Chandigarh', 'Chennai', 'Delhi', 'Hyderabad','Indore', 'Jaipur', 'Mumbai', 'Pune', 'Surat', 'In-Organic','Total']]
    aum_product_trend=aum_product_trend.transpose()
    ############################ AUM PRODUCT TREND COMPLETED#######################
    
    ############################ TRANCH DISBURSEMENT #############################
    def disb_tranch(disb,eomfin,t):
        if t==1:
            disb=disb[disb['DISBURSEMENT_SEQUENCE']==1]
        elif t==0:
            disb=disb[disb['DISBURSEMENT_SEQUENCE']!=1]
        else:
            disb=disb
        
        disbursemnt={}
        
        # timedelta() gets successive dates with
        # appropriate difference
        dates_2020 = [ datetime.date(2020, 4, 1) + datetime.timedelta(days=idx) for idx in range(365)]
        dates_2021= [ datetime.date(2021, 4, 1) + datetime.timedelta(days=idx) for idx in range(365)]
        
        year=[]
        for i in ytd:
            year.append([i])
        for j in year:
            eomdisb_1=disb.copy()
            eomdisb=pd.merge(eomdisb_1,eomfin,on=('LAN_ID'),how='left')
            eomdisb=eomdisb[eomdisb['DISB_YEAR_MONTH'].isin(j)]
            # eomdisb['DISB_VOL']=np.where((eomdisb['DISB_YEAR_MONTH'].isin(MTD)),1,0)
            eomdisb['DISB_VOL']=np.where(True,1,0)
            eomdisb['DISBURSEMENT_AMOUNT']=eomdisb['DISBURSEMENT_AMOUNT']/1000000000
            eomdisb['DISBURSEMENT_AMOUNT']=eomdisb['DISBURSEMENT_AMOUNT'].round(2)
            eomdisb=eomdisb[[ 'REPORTING_BRANCH', 'DISBURSEMENT_AMOUNT','FINTYPE']]
            eomdisbview=eomdisb.groupby('REPORTING_BRANCH').sum(['DISBURSEMENT_AMOUNT'])
            monbr=list(set(list(base_view.REPORTING_BRANCH.unique()))-set(eomdisbview.index.to_list()))
            for z in monbr:
                eomdisbview.loc[z,:]=0
            eomdisbview=eomdisbview.sort_values(by='REPORTING_BRANCH')
            eomdisbview.loc['Total']= eomdisbview.sum()
            eomdisbview.rename(columns={'DISBURSEMENT_AMOUNT':j[0]},inplace=True)
            disbursemnt[j[0]]=eomdisbview
        eomdisb_1=disb.copy()
        eomdisb=pd.merge(eomdisb_1,eomfin,on=('LAN_ID'),how='left')
        eomdisb['DISBDATE']=np.where(True,pd.to_datetime(eomdisb['DISBURSEMENT_DATE']).dt.date ,0)
        eomdisb=eomdisb[eomdisb['DISBDATE'].isin(dates_2020)]
        # eomdisb['DISB_VOL']=np.where((eomdisb['DISB_YEAR_MONTH'].isin(MTD)),1,0)
        eomdisb['DISB_VOL']=np.where(True,1,0)
        eomdisb['DISBURSEMENT_AMOUNT']=eomdisb['DISBURSEMENT_AMOUNT']/1000000000
        eomdisb['DISBURSEMENT_AMOUNT']=eomdisb['DISBURSEMENT_AMOUNT'].round(2)
        eomdisb=eomdisb[[ 'REPORTING_BRANCH', 'DISBURSEMENT_AMOUNT','FINTYPE']]
        eomdisbview=eomdisb.groupby('REPORTING_BRANCH').sum(['DISBURSEMENT_AMOUNT'])
        monbr=list(set(list(base_view.REPORTING_BRANCH.unique()))-set(eomdisbview.index.to_list()))
        for z in monbr:
            eomdisbview.loc[z,:]=0
        eomdisbview=eomdisbview.sort_values(by='REPORTING_BRANCH')
        eomdisbview.loc['Total']= eomdisbview.sum()
        eomdisbview.rename(columns={'DISBURSEMENT_AMOUNT':2020},inplace=True)
        disbursemnt['2020']=eomdisbview
        
        eomdisb_1=disb.copy()
        eomdisb=pd.merge(eomdisb_1,eomfin,on=('LAN_ID'),how='left')
        eomdisb['DISBDATE']=np.where(True,pd.to_datetime(eomdisb['DISBURSEMENT_DATE']).dt.date ,0)
        eomdisb=eomdisb[eomdisb['DISBDATE'].isin(dates_2021)]
        # eomdisb['DISB_VOL']=np.where((eomdisb['DISB_YEAR_MONTH'].isin(MTD)),1,0)
        eomdisb['DISB_VOL']=np.where(True,1,0)
        eomdisb['DISBURSEMENT_AMOUNT']=eomdisb['DISBURSEMENT_AMOUNT']/1000000000
        eomdisb['DISBURSEMENT_AMOUNT']=eomdisb['DISBURSEMENT_AMOUNT'].round(2)
        eomdisb=eomdisb[[ 'REPORTING_BRANCH', 'DISBURSEMENT_AMOUNT','FINTYPE']]
        eomdisbview=eomdisb.groupby('REPORTING_BRANCH').sum(['DISBURSEMENT_AMOUNT'])
        monbr=list(set(list(base_view.REPORTING_BRANCH.unique()))-set(eomdisbview.index.to_list()))
        for z in monbr:
            eomdisbview.loc[z,:]=0
        eomdisbview=eomdisbview.sort_values(by='REPORTING_BRANCH')
        eomdisbview.loc['Total']= eomdisbview.sum()
        eomdisbview.rename(columns={'DISBURSEMENT_AMOUNT':2021},inplace=True)
        disbursemnt['2021']=eomdisbview
        
        dfs=disbursemnt.values()
        merged_df = reduce(lambda l, r: pd.merge(l, r, on='REPORTING_BRANCH', how='inner'), dfs)
        total_disb=merged_df.round(2)
        total_disb=total_disb[[      2020,      2021,  '2022-4',  '2022-5',  '2022-6',  '2022-7','2022-8',  '2022-9', '2022-10', '2022-11', '2022-12']]
        return total_disb
    tranch1=disb_tranch(disb,eomfin,1)
    tranch0=disb_tranch(disb,eomfin,0)
    tranchall=disb_tranch(disb,eomfin,2)
    
    ############################ TRANCH DISBURSEMENT COMPLETE #############################
    
    
    
    ############################ FTD #############################################
    
    '''
    HL
    '''
    hl=eomfin.copy()
    hl_disb=eomdisb_1.copy()
    hl=hl[hl['FINTYPE'].isin(['HL'])]
    hlf,hlfp=end_of(day,hl,hl_disb,LRD='LRD',Rd=True)
    hlfvol=hlf[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    hlfval=hlf[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    hlfval['DISB']=hlfval['DISBURSEMENT_AMOUNT_TRANCH1']+hlfval['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    hlfval=hlfval[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    
    bt=eomfin.copy()
    bt_disb=eomdisb_1.copy()
    bt=bt[bt['FINTYPE'].isin(['HT','LT'])]
    btf,btfp=end_of(day,bt,bt_disb,LRD='LRD',Rd=True)
    btfvol=btf[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    btfval=btf[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    btfval['DISB']=btfval['DISBURSEMENT_AMOUNT_TRANCH1']+btfval['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    btfval=btfval[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    fl=eomfin.copy()
    fl_disb=eomdisb_1.copy()
    fl=fl[fl['FINTYPE'].isin(['FL','FT'])]
    flf,flfp=end_of(day,fl,fl_disb,LRD='LRD',Rd=True)
    flfvol=flf[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    flfval=flf[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    flfval['DISB']=flfval['DISBURSEMENT_AMOUNT_TRANCH1']+flfval['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    flfval=flfval[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    '''
    LAP
    '''
    nlap=eomfin.copy()
    nlap_disb=eomdisb_1.copy()
    nlap=nlap[(nlap['FINTYPE'].isin(['LP','NP']))&(nlap['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP']))]
    nlapf,nlapmp=end_of(day,nlap,nlap_disb,LRD='LRD',Rd=True)
    nlapf_vol=nlapf[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    nlapf_val=nlapf[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    nlapf_val['DISB']=nlapf_val['DISBURSEMENT_AMOUNT_TRANCH1']+nlapf_val['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    nlapf_val=nlapf_val[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    
    lrd=eomfin.copy()
    lrd_disb=eomdisb_1.copy()
    lrd=lrd[(lrd['FINTYPE'].isin(['LP','NP']))&(lrd['SUB_PRODUCT'].isin(['LRD']))]
    lrdf,lrdfp=end_of(day,lrd,lrd_disb,LRD='LRD',Rd=True)
    lrdf_vol=lrdf[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    lrdf_val=lrdf[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    lrdf_val['DISB']=lrdf_val['DISBURSEMENT_AMOUNT_TRANCH1']+lrdf_val['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    lrdf_val=lrdf_val[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    lap=eomfin.copy()
    lap_disb=eomdisb_1.copy()
    lap=lap[(lap['FINTYPE'].isin(['LP']))&(lap['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP','LRD'])==False)]
    lapf,lapfp=end_of(day,lap,lap_disb,LRD='LRD',Rd=True)
    lapf_vol=lapf[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    lapf_val=lapf[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    lapf_val['DISB']=lapf_val['DISBURSEMENT_AMOUNT_TRANCH1']+lapf_val['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    lapf_val=lapf_val[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    nrp=eomfin.copy()
    nrp_disb=eomdisb_1.copy()
    nrp=nrp[(nrp['FINTYPE'].isin(['NP']))&(nrp['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP','LRD'])==False)]
    nrpf,nrpfp=end_of(day,nrp,nrp_disb,LRD='LRD',Rd=True)
    nrpf_vol=nrpf[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    nrpf_val=nrpf[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    nrpf_val['DISB']=nrpf_val['DISBURSEMENT_AMOUNT_TRANCH1']+nrpf_val['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    nrpf_val=nrpf_val[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    ############################ MTD #############################################
    
    '''
    HL
    '''
    hl=eomfin.copy()
    hl_disb=eomdisb_1.copy()
    hl=hl[hl['FINTYPE'].isin(['HL'])]
    hlm,hlmp=end_of(MTD,hl,hl_disb,LRD='LRD',Rd=True)
    hlmvol=hlm[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    hlmval=hlm[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    hlmval['DISB']=hlmval['DISBURSEMENT_AMOUNT_TRANCH1']+hlmval['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    hlmval=hlmval[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    bt=eomfin.copy()
    bt_disb=eomdisb_1.copy()
    bt=bt[bt['FINTYPE'].isin(['HT','LT'])]
    btm,btfm=end_of(MTD,bt,bt_disb,LRD='LRD',Rd=True)
    btmvol=btm[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    btmval=btm[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    btmval['DISB']=btmval['DISBURSEMENT_AMOUNT_TRANCH1']+btmval['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    btmval=btmval[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    fl=eomfin.copy()
    fl_disb=eomdisb_1.copy()
    fl=fl[fl['FINTYPE'].isin(['FL','FT'])]
    flm,flmp=end_of(MTD,fl,fl_disb,LRD='LRD',Rd=True)
    flmvol=flm[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    flmval=flm[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    flmval['DISB']=flmval['DISBURSEMENT_AMOUNT_TRANCH1']+flmval['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    flmval=flmval[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    '''
    LAP
    '''
    nlap=eomfin.copy()
    nlap_disb=eomdisb_1.copy()
    nlap=nlap[(nlap['FINTYPE'].isin(['LP','NP']))&(nlap['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP']))]
    nlapm,nlapmp=end_of(MTD,nlap,nlap_disb,LRD='LRD',Rd=True)
    nlapm_vol=nlapm[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    nlapm_val=nlapm[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    nlapm_val['DISB']=nlapm_val['DISBURSEMENT_AMOUNT_TRANCH1']+nlapm_val['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    nlapm_val=nlapm_val[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    
    lrd=eomfin.copy()
    lrd_disb=eomdisb_1.copy()
    lrd=lrd[(lrd['FINTYPE'].isin(['LP','NP']))&(lrd['SUB_PRODUCT'].isin(['LRD']))]
    lrdm,lrdmp=end_of(MTD,lrd,lrd_disb,LRD='LRD',Rd=True)
    lrdm_vol=lrdm[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    lrdm_val=lrdm[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    lrdm_val['DISB']=lrdm_val['DISBURSEMENT_AMOUNT_TRANCH1']+lrdm_val['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    lrdm_val=lrdm_val[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    lap=eomfin.copy()
    lap_disb=eomdisb_1.copy()
    lap=lap[(lap['FINTYPE'].isin(['LP']))&(lap['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP','LRD'])==False)]
    lapm,lapmp=end_of(MTD,lap,lap_disb,LRD='LRD',Rd=True)
    lapm_vol=lapm[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    lapm_val=lapm[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    lapm_val['DISB']=lapm_val['DISBURSEMENT_AMOUNT_TRANCH1']+lapm_val['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    lapm_val=lapm_val[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    nrp=eomfin.copy()
    nrp_disb=eomdisb_1.copy()
    nrp=nrp[(nrp['FINTYPE'].isin(['NP']))&(nrp['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP','LRD'])==False)]
    nrpm,nrpmp=end_of(MTD,nrp,nrp_disb,LRD='LRD',Rd=True)
    nrpm_vol=nrpm[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    nrpm_val=nrpm[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    nrpm_val['DISB']=nrpm_val['DISBURSEMENT_AMOUNT_TRANCH1']+nrpm_val['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    nrpm_val=nrpm_val[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    
    ############################ YTD #############################################
    
    '''
    HL
    '''
    hl=eomfin.copy()
    hl_disb=eomdisb_1.copy()
    hl=hl[hl['FINTYPE'].isin(['HL'])]
    hly,hlyp=end_of(ytd,hl,hl_disb,LRD='LRD',Rd=True)
    hlyvol=hly[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    hlyval=hly[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    hlyval['DISB']=hlyval['DISBURSEMENT_AMOUNT_TRANCH1']+hlyval['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    hlyval=hlyval[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    bt=eomfin.copy()
    bt_disb=eomdisb_1.copy()
    bt=bt[bt['FINTYPE'].isin(['HT','LT'])]
    bty,btfy=end_of(ytd,bt,bt_disb,LRD='LRD',Rd=True)
    btyvol=bty[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    btyval=bty[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    btyval['DISB']=btyval['DISBURSEMENT_AMOUNT_TRANCH1']+btyval['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    btyval=btyval[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    fl=eomfin.copy()
    fl_disb=eomdisb_1.copy()
    fl=fl[fl['FINTYPE'].isin(['FL','FT'])]
    fly,flyp=end_of(ytd,fl,fl_disb,LRD='LRD',Rd=True)
    flyvol=fly[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    flyval=fly[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    flyval['DISB']=flyval['DISBURSEMENT_AMOUNT_TRANCH1']+flyval['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    flyval=flyval[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    '''
    LAP
    '''
    nlap=eomfin.copy()
    nlap_disb=eomdisb_1.copy()
    nlap=nlap[(nlap['FINTYPE'].isin(['LP','NP']))&(nlap['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP']))]
    nlapy,nlapyp=end_of(ytd,nlap,nlap_disb,LRD='LRD',Rd=True)
    nlapy_vol=nlapy[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    nlapy_val=nlapy[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    nlapy_val['DISB']=nlapy_val['DISBURSEMENT_AMOUNT_TRANCH1']+nlapy_val['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    nlapy_val=nlapy_val[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    
    lrd=eomfin.copy()
    lrd_disb=eomdisb_1.copy()
    lrd=lrd[(lrd['FINTYPE'].isin(['LP','NP']))&(lrd['SUB_PRODUCT'].isin(['LRD']))]
    lrdy,lrdyp=end_of(ytd,lrd,lrd_disb,LRD='LRD',Rd=True)
    lrdy_vol=lrdy[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    lrdy_val=lrdy[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    lrdy_val['DISB']=lrdy_val['DISBURSEMENT_AMOUNT_TRANCH1']+lrdy_val['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    lrdy_val=lrdy_val[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    lap=eomfin.copy()
    lap_disb=eomdisb_1.copy()
    lap=lap[(lap['FINTYPE'].isin(['LP']))&(lap['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP','LRD'])==False)]
    lapy,lapyp=end_of(ytd,lap,lap_disb,LRD='LRD',Rd=True)
    lapy_vol=lapy[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    lapy_val=lapy[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    lapy_val['DISB']=lapy_val['DISBURSEMENT_AMOUNT_TRANCH1']+lapy_val['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    lapy_val=lapy_val[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    nrp=eomfin.copy()
    nrp_disb=eomdisb_1.copy()
    nrp=nrp[(nrp['FINTYPE'].isin(['NP']))&(nrp['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP','LRD'])==False)]
    nrpy,nrpyp=end_of(ytd,nrp,nrp_disb,LRD='LRD',Rd=True)
    nrpy_vol=nrpy[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    nrpy_val=nrpy[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    nrpy_val['DISB']=nrpy_val['DISBURSEMENT_AMOUNT_TRANCH1']+nrpy_val['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    nrpy_val=nrpy_val[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    
    ############################ MTD Funnel#############################################
    
    '''
    HL
    '''
    hl=eomfin.copy()
    hl_disb=eomdisb_1.copy()
    hl=hl[hl['FINTYPE'].isin(['HL'])]
    hl=hl[hl['LOGIN_YEAR_MONTH'].isin(MTD)]
    hlmf,hlmfp=end_of(MTD,hl,hl_disb,LRD='LRD')
    hlmfvol=hlmf[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    hlmfval=hlmf[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    hlmfval['DISB']=hlmfval['DISBURSEMENT_AMOUNT_TRANCH1']+hlmfval['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    hlmfval=hlmfval[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    bt=eomfin.copy()
    bt_disb=eomdisb_1.copy()
    bt=bt[bt['FINTYPE'].isin(['HT','LT'])]
    bt=bt[bt['LOGIN_YEAR_MONTH'].isin(MTD)]
    btmf,btfmf=end_of(MTD,bt,bt_disb,LRD='LRD')
    btmfvol=btmf[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    btmfval=btmf[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    btmfval['DISB']=btmfval['DISBURSEMENT_AMOUNT_TRANCH1']+btmfval['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    btmfval=btmfval[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    fl=eomfin.copy()
    fl_disb=eomdisb_1.copy()
    fl=fl[fl['FINTYPE'].isin(['FL','FT'])]
    fl=fl[fl['LOGIN_YEAR_MONTH'].isin(MTD)]
    flmf,flmfp=end_of(MTD,fl,fl_disb,LRD='LRD')
    flmfvol=flmf[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    flmfval=flmf[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    flmfval['DISB']=flmfval['DISBURSEMENT_AMOUNT_TRANCH1']+flmfval['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    flmfval=flmfval[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    '''
    LAP
    '''
    nlap=eomfin.copy()
    nlap_disb=eomdisb_1.copy()
    nlap=nlap[(nlap['FINTYPE'].isin(['LP','NP']))&(nlap['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP']))]
    nlap=nlap[nlap['LOGIN_YEAR_MONTH'].isin(MTD)]
    nlapmf,nlapmpf=end_of(MTD,nlap,nlap_disb,LRD='LRD')
    nlapmf_vol=nlapmf[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    nlapmf_val=nlapmf[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    nlapmf_val['DISB']=nlapmf_val['DISBURSEMENT_AMOUNT_TRANCH1']+nlapmf_val['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    nlapmf_val=nlapmf_val[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    
    lrd=eomfin.copy()
    lrd_disb=eomdisb_1.copy()
    lrd=lrd[(lrd['FINTYPE'].isin(['LP','NP']))&(lrd['SUB_PRODUCT'].isin(['LRD']))]
    lrd=lrd[lrd['LOGIN_YEAR_MONTH'].isin(MTD)]
    lrdmf,lrdmfp=end_of(MTD,lrd,lrd_disb,LRD='LRD')
    lrdmf_vol=lrdmf[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    lrdmf_val=lrdmf[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    lrdmf_val['DISB']=lrdmf_val['DISBURSEMENT_AMOUNT_TRANCH1']+lrdmf_val['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    lrdmf_val=lrdmf_val[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    lap=eomfin.copy()
    lap_disb=eomdisb_1.copy()
    lap=lap[(lap['FINTYPE'].isin(['LP']))&(lap['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP','LRD'])==False)]
    lap=lap[lap['LOGIN_YEAR_MONTH'].isin(MTD)]
    lapmf,lapmpf=end_of(MTD,lap,lap_disb,LRD='LRD')
    lapmf_vol=lapmf[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    lapmf_val=lapmf[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    lapmf_val['DISB']=lapmf_val['DISBURSEMENT_AMOUNT_TRANCH1']+lapmf_val['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    lapmf_val=lapmf_val[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    
    nrp=eomfin.copy()
    nrp_disb=eomdisb_1.copy()
    nrp=nrp[(nrp['FINTYPE'].isin(['NP']))&(nrp['SUB_PRODUCT'].isin(['NEO LAP','NEO NRP','NEO BOOSTER LAP','LRD'])==False)]
    nrp=nrp[nrp['LOGIN_YEAR_MONTH'].isin(MTD)]
    nrpmf,nrpmfp=end_of(MTD,nrp,nrp_disb,LRD='LRD')
    nrpmf_vol=nrpmf[['LOG_VOL', 'IN_SANCTION_VOL', 'SANCTION_VOL','BOOK_VOL','REJECT_VOL']]
    nrpmf_val=nrpmf[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES', 'Premium','REJECT_AMOUNT']]
    
    nrpmf_val['DISB']=nrpmf_val['DISBURSEMENT_AMOUNT_TRANCH1']+nrpmf_val['DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES']
    nrpmf_val=nrpmf_val[['REQUESTED_AMOUNT','IN_SANCTION_AMOUNT', 'SANCTION_AMOUNT','BOOKING_AMOUNT','DISB', 'Premium','REJECT_AMOUNT']]
    
    ###############################################################################
    
    
    
    
    
    
    
    ############################ WRITING DATA IN EXCEL ##############################
    
    from openpyxl import load_workbook
    import os
    import time
    # current_dir =r"\\GHFL-SNOWFLAKES\Users\bischeduler\Documents\Output_of_Scheduler\EOM_REPORTS"
    current_dir =r"\\GHFL-SNOWFLAKES\Users\bischeduler\Documents\Output_of_Scheduler\BI_Daily_Reports"
    os.makedirs(f'{current_dir}/{time.strftime("/%Y-%m-%d")}', exist_ok=True)
    mtd_df=mtd_df[['LOG_VOL', 'REQUESTED_AMOUNT', 'SANCTION_VOL','SANCTION_AMOUNT', 'BOOK_VOL', 'BOOKING_AMOUNT', 'Disbursement_Volume','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES','Premium','Penetration','ROI']]
    ytd_df=ytd_df[['LOG_VOL', 'REQUESTED_AMOUNT', 'SANCTION_VOL','SANCTION_AMOUNT', 'BOOK_VOL', 'BOOKING_AMOUNT', 'Disbursement_Volume','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES','Premium','Penetration','ROI']]
    ftd_df=ftd_df[['LOG_VOL', 'REQUESTED_AMOUNT', 'SANCTION_VOL','SANCTION_AMOUNT', 'BOOK_VOL', 'BOOKING_AMOUNT', 'Disbursement_Volume','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES','Premium','Penetration','ROI']]
    eomp=eomp[['LOG_VOL', 'REQUESTED_AMOUNT', 'SANCTION_VOL','SANCTION_AMOUNT', 'BOOK_VOL', 'BOOKING_AMOUNT', 'Disbursement_Volume','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES','Premium','Penetration','ROI']]
    eodp=eodp[['LOG_VOL', 'REQUESTED_AMOUNT', 'SANCTION_VOL','SANCTION_AMOUNT', 'BOOK_VOL', 'BOOKING_AMOUNT', 'Disbursement_Volume','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES','Premium','Penetration','ROI']]
    eoyp=eoyp[['LOG_VOL', 'REQUESTED_AMOUNT', 'SANCTION_VOL','SANCTION_AMOUNT', 'BOOK_VOL', 'BOOKING_AMOUNT', 'Disbursement_Volume','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES','Premium','Penetration','ROI']]
    stlm_df=stlm_df[['LOG_VOL', 'REQUESTED_AMOUNT', 'SANCTION_VOL','SANCTION_AMOUNT', 'BOOK_VOL', 'BOOKING_AMOUNT', 'Disbursement_Volume','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES','Premium','Penetration','ROI']]
    stlmp=stlmp[['LOG_VOL', 'REQUESTED_AMOUNT', 'SANCTION_VOL','SANCTION_AMOUNT', 'BOOK_VOL', 'BOOKING_AMOUNT', 'Disbursement_Volume','DISBURSEMENT_AMOUNT_TRANCH1','DISBURSEMENT_AMOUNT_SUBSEQUENT_TRANCHES','Premium','Penetration','ROI']]
    
    
    mtd_df.reset_index(inplace=True)
    mtd_df.rename(columns={'REPORTING_BRANCH':" "},inplace=True)
    ytd_df.reset_index(inplace=True)
    ytd_df.rename(columns={'REPORTING_BRANCH':" "},inplace=True)
    ftd_df.reset_index(inplace=True)
    ftd_df.rename(columns={'REPORTING_BRANCH':" "},inplace=True)
    stlm_df.reset_index(inplace=True)
    stlm_df.rename(columns={'REPORTING_BRANCH':" "},inplace=True)
    
    
    eomp.reset_index(inplace=True)
    eomp.rename(columns={'Prod':" "},inplace=True)
    eodp.reset_index(inplace=True)
    eodp.rename(columns={'Prod':" "},inplace=True)
    eoyp.reset_index(inplace=True)
    eoyp.rename(columns={'Prod':" "},inplace=True)
    stlmp.reset_index(inplace=True)
    stlmp.rename(columns={'Prod':" "},inplace=True)
    
    
    
    os.chdir(f'{current_dir}')
    fn = "losnewformat4.xlsx"
    book = load_workbook(fn)
    os.chdir(f'{current_dir}/{time.strftime("/%Y-%m-%d")}')
    
    writer = pd.ExcelWriter("LOS_GC_SUMMARY1.xlsx", engine='openpyxl')
    
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    
    ################################SUMMARY########################################
    
    ftd_df.to_excel(writer, sheet_name='Summary', header=None, index=False,
                 startcol=2, startrow=49,)
    mtd_df.to_excel(writer, sheet_name='Summary', header=None, index=False,
                 startcol=16, startrow=49)
    ytd_df.to_excel(writer, sheet_name='Summary', header=None, index=False,
                 startcol=2, startrow=92)
    stlm_df.to_excel(writer, sheet_name='Summary', header=None, index=False,
                 startcol=16, startrow=92)
    
    
    eodp.to_excel(writer, sheet_name='Summary', header=None, index=False,
                 startcol=2, startrow=27
                 )
    eomp.to_excel(writer, sheet_name='Summary', header=None, index=False,
                 startcol=16, startrow=27)
    eoyp.to_excel(writer, sheet_name='Summary', header=None, index=False,
                 startcol=2, startrow=37)
    stlmp.to_excel(writer, sheet_name='Summary', header=None, index=False,
                 startcol=16, startrow=37)
    aum['INORG']=' '
    aum['INORGANIC_TARGETS']=' '
    aum.to_excel(writer, sheet_name='Summary', header=None, index=False,startcol=2, startrow=7)
    
    
    
    
    
    #######################SUMMARY DONE ###########################################
    
    ftd_val.to_excel(writer, sheet_name='Branch Dashboard', header=None, index=False,startcol=2, startrow=6)
    ytd_val.to_excel(writer, sheet_name='Branch Dashboard', header=None, index=False,startcol=2, startrow=21)
    mtd_val.to_excel(writer, sheet_name='Branch Dashboard', header=None, index=False,startcol=10, startrow=6)
    stlm_val.to_excel(writer, sheet_name='Branch Dashboard', header=None, index=False,startcol=10, startrow=21)
    
    ftd_vol.to_excel(writer, sheet_name='Branch Dashboard', header=None, index=False,startcol=2, startrow=38)
    ytd_vol.to_excel(writer, sheet_name='Branch Dashboard', header=None, index=False,startcol=2, startrow=53)
    mtd_vol.to_excel(writer, sheet_name='Branch Dashboard', header=None, index=False,startcol=10, startrow=38)
    stlm_vol.to_excel(writer, sheet_name='Branch Dashboard', header=None, index=False,startcol=10, startrow=53)
    adoption.to_excel(writer, sheet_name='Branch Dashboard', header=None, index=False,startcol=2, startrow=71)
    ###############################################################################
    ################################AUM PRODUCT TREND##############################
    aum_product_trend.reset_index(inplace=True )
    aum_product_trend.to_excel(writer, sheet_name='AUM_Product_Trend', header=None, index=False,startcol=3, startrow=4)
    #####################AUM PRODUCT TREND DONE###################################
    ##################### INCOME SANCTION FLOW ####################################
    sanc_val.reset_index(inplace=True)
    sanc_vol.reset_index(inplace=True)
    sanc_val.to_excel(writer, sheet_name='Income Sanction flow', header=None, index=False,startcol=3, startrow=6)
    sanc_vol.to_excel(writer, sheet_name='Income Sanction flow', header=None, index=False,startcol=27, startrow=6)
    ##############################################################################
    
    ############################ FTD ############################################
    
    
    hlfval.reset_index(inplace=True)
    hlfvol.reset_index(inplace=True)
    btfval.reset_index(inplace=True)
    btfvol.reset_index(inplace=True)
    flfval.reset_index(inplace=True)
    flfvol.reset_index(inplace=True)
    nlapf_val.reset_index(inplace=True)
    nlapf_vol.reset_index(inplace=True)
    lrdf_val.reset_index(inplace=True)
    lrdf_vol.reset_index(inplace=True)
    lapf_val.reset_index(inplace=True)
    lapf_vol.reset_index(inplace=True)
    nrpf_vol.reset_index(inplace=True)
    nrpf_val.reset_index(inplace=True)
    
    
    hlfval.to_excel(writer, sheet_name='FTD', header=None, index=False,startcol=10, startrow=8)
    hlfvol.to_excel(writer, sheet_name='FTD', header=None, index=False,startcol=3, startrow=8)
    btfval.to_excel(writer, sheet_name='FTD', header=None, index=False,startcol=10, startrow=23)
    btfvol.to_excel(writer, sheet_name='FTD', header=None, index=False,startcol=3, startrow=23)
    flfval.to_excel(writer, sheet_name='FTD', header=None, index=False,startcol=10, startrow=38)
    flfvol.to_excel(writer, sheet_name='FTD', header=None, index=False,startcol=3, startrow=38)
    nlapf_val.to_excel(writer, sheet_name='FTD', header=None, index=False,startcol=10, startrow=55)
    nlapf_vol.to_excel(writer, sheet_name='FTD', header=None, index=False,startcol=3, startrow=55)
    lrdf_val.to_excel(writer, sheet_name='FTD', header=None, index=False,startcol=10, startrow=70)
    lrdf_vol.to_excel(writer, sheet_name='FTD', header=None, index=False,startcol=3, startrow=70)
    lapf_val.to_excel(writer, sheet_name='FTD', header=None, index=False,startcol=10, startrow=85)
    lapf_vol.to_excel(writer, sheet_name='FTD', header=None, index=False,startcol=3, startrow=85)
    nrpf_vol.to_excel(writer, sheet_name='FTD', header=None, index=False,startcol=3, startrow=100)
    nrpf_val.to_excel(writer, sheet_name='FTD', header=None, index=False,startcol=10, startrow=100)
    
    ############################# MTD ############################################
    
    hlmval.reset_index(inplace=True)
    hlmvol.reset_index(inplace=True)
    btmval.reset_index(inplace=True)
    btmvol.reset_index(inplace=True)
    flmval.reset_index(inplace=True)
    flmvol.reset_index(inplace=True)
    nlapm_val.reset_index(inplace=True)
    nlapm_vol.reset_index(inplace=True)
    lrdm_val.reset_index(inplace=True)
    lrdm_vol.reset_index(inplace=True)
    lapm_val.reset_index(inplace=True)
    lapm_vol.reset_index(inplace=True)
    nrpm_vol.reset_index(inplace=True)
    nrpm_val.reset_index(inplace=True)
    
    hlmval.to_excel(writer, sheet_name='MTD', header=None, index=False,startcol=10, startrow=8)
    hlmvol.to_excel(writer, sheet_name='MTD', header=None, index=False,startcol=3, startrow=8)
    btmval.to_excel(writer, sheet_name='MTD', header=None, index=False,startcol=10, startrow=23)
    btmvol.to_excel(writer, sheet_name='MTD', header=None, index=False,startcol=3, startrow=23)
    flmval.to_excel(writer, sheet_name='MTD', header=None, index=False,startcol=10, startrow=38)
    flmvol.to_excel(writer, sheet_name='MTD', header=None, index=False,startcol=3, startrow=38)
    nlapm_val.to_excel(writer, sheet_name='MTD', header=None, index=False,startcol=10, startrow=55)
    nlapm_vol.to_excel(writer, sheet_name='MTD', header=None, index=False,startcol=3, startrow=55)
    lrdm_val.to_excel(writer, sheet_name='MTD', header=None, index=False,startcol=10, startrow=70)
    lrdm_vol.to_excel(writer, sheet_name='MTD', header=None, index=False,startcol=3, startrow=70)
    lapm_val.to_excel(writer, sheet_name='MTD', header=None, index=False,startcol=10, startrow=85)
    lapm_vol.to_excel(writer, sheet_name='MTD', header=None, index=False,startcol=3, startrow=85)
    nrpm_vol.to_excel(writer, sheet_name='MTD', header=None, index=False,startcol=3, startrow=100)
    nrpm_val.to_excel(writer, sheet_name='MTD', header=None, index=False,startcol=10, startrow=100)
    
    ############################# YTD ############################################
    hlyval.reset_index(inplace=True)
    hlyvol.reset_index(inplace=True)
    btyval.reset_index(inplace=True)
    btyvol.reset_index(inplace=True)
    flyval.reset_index(inplace=True)
    flyvol.reset_index(inplace=True)
    nlapy_val.reset_index(inplace=True)
    nlapy_vol.reset_index(inplace=True)
    lrdy_val.reset_index(inplace=True)
    lrdy_vol.reset_index(inplace=True)
    lapy_val.reset_index(inplace=True)
    lapy_vol.reset_index(inplace=True)
    nrpy_vol.reset_index(inplace=True)
    nrpy_val.reset_index(inplace=True)
    
    hlyval.to_excel(writer, sheet_name='YTD', header=None, index=False,startcol=10, startrow=8)
    hlyvol.to_excel(writer, sheet_name='YTD', header=None, index=False,startcol=3, startrow=8)
    btyval.to_excel(writer, sheet_name='YTD', header=None, index=False,startcol=10, startrow=23)
    btyvol.to_excel(writer, sheet_name='YTD', header=None, index=False,startcol=3, startrow=23)
    flyval.to_excel(writer, sheet_name='YTD', header=None, index=False,startcol=10, startrow=38)
    flyvol.to_excel(writer, sheet_name='YTD', header=None, index=False,startcol=3, startrow=38)
    nlapy_val.to_excel(writer, sheet_name='YTD', header=None, index=False,startcol=10, startrow=55)
    nlapy_vol.to_excel(writer, sheet_name='YTD', header=None, index=False,startcol=3, startrow=55)
    lrdy_val.to_excel(writer, sheet_name='YTD', header=None, index=False,startcol=10, startrow=70)
    lrdy_vol.to_excel(writer, sheet_name='YTD', header=None, index=False,startcol=3, startrow=70)
    lapy_val.to_excel(writer, sheet_name='YTD', header=None, index=False,startcol=10, startrow=85)
    lapy_vol.to_excel(writer, sheet_name='YTD', header=None, index=False,startcol=3, startrow=85)
    nrpy_vol.to_excel(writer, sheet_name='YTD', header=None, index=False,startcol=3, startrow=100)
    nrpy_val.to_excel(writer, sheet_name='YTD', header=None, index=False,startcol=10, startrow=100)
    
    
    ############################# MTD Funnel ######################################
    hlmfval.reset_index(inplace=True)
    hlmfvol.reset_index(inplace=True)
    btmfval.reset_index(inplace=True)
    btmfvol.reset_index(inplace=True)
    flmfval.reset_index(inplace=True)
    flmfvol.reset_index(inplace=True)
    nlapmf_val.reset_index(inplace=True)
    nlapmf_vol.reset_index(inplace=True)
    lrdmf_val.reset_index(inplace=True)
    lrdmf_vol.reset_index(inplace=True)
    lapmf_val.reset_index(inplace=True)
    lapmf_vol.reset_index(inplace=True)
    nrpmf_vol.reset_index(inplace=True)
    nrpmf_val.reset_index(inplace=True)
    
    hlmfval.to_excel(writer, sheet_name='MTD Funnel', header=None, index=False,startcol=10, startrow=8)
    hlmfvol.to_excel(writer, sheet_name='MTD Funnel', header=None, index=False,startcol=3, startrow=8)
    btmfval.to_excel(writer, sheet_name='MTD Funnel', header=None, index=False,startcol=10, startrow=23)
    btmfvol.to_excel(writer, sheet_name='MTD Funnel', header=None, index=False,startcol=3, startrow=23)
    flmfval.to_excel(writer, sheet_name='MTD Funnel', header=None, index=False,startcol=10, startrow=38)
    flmfvol.to_excel(writer, sheet_name='MTD Funnel', header=None, index=False,startcol=3, startrow=38)
    nlapmf_val.to_excel(writer, sheet_name='MTD Funnel', header=None, index=False,startcol=10, startrow=55)
    nlapmf_vol.to_excel(writer, sheet_name='MTD Funnel', header=None, index=False,startcol=3, startrow=55)
    lrdmf_val.to_excel(writer, sheet_name='MTD Funnel', header=None, index=False,startcol=10, startrow=70)
    lrdmf_vol.to_excel(writer, sheet_name='MTD Funnel', header=None, index=False,startcol=3, startrow=70)
    lapmf_val.to_excel(writer, sheet_name='MTD Funnel', header=None, index=False,startcol=10, startrow=85)
    lapmf_vol.to_excel(writer, sheet_name='MTD Funnel', header=None, index=False,startcol=3, startrow=85)
    nrpmf_vol.to_excel(writer, sheet_name='MTD Funnel', header=None, index=False,startcol=3, startrow=100)
    nrpmf_val.to_excel(writer, sheet_name='MTD Funnel', header=None, index=False,startcol=10, startrow=100)
    
    ################################# Tranch Disb #############################
    tranch0.reset_index(inplace=True)
    tranch1.reset_index(inplace=True)
    tranchall.reset_index(inplace=True)
    tranch0.to_excel(writer, sheet_name='Tranche disbursement', header=None, index=False,startcol=17, startrow=23)
    tranch1.to_excel(writer, sheet_name='Tranche disbursement', header=None, index=False,startcol=3, startrow=23)
    tranchall.to_excel(writer, sheet_name='Tranche disbursement', header=None, index=False,startcol=3, startrow=6)
    
    ################################ PRODUCT DASHBOARD ############################
    
    pmvol.to_excel(writer, sheet_name='Product Dashboard', header=None, index=False,startcol=10, startrow=34,)
    pmval.to_excel(writer, sheet_name='Product Dashboard', header=None, index=False,startcol=10, startrow=6)
    pyvol.to_excel(writer, sheet_name='Product Dashboard', header=None, index=False,startcol=2, startrow=47)
    pyval.to_excel(writer, sheet_name='Product Dashboard', header=None, index=False,startcol=2, startrow=19)
    pdvol.to_excel(writer, sheet_name='Product Dashboard', header=None, index=False,startcol=2, startrow=34)
    pdval.to_excel(writer, sheet_name='Product Dashboard', header=None, index=False,startcol=2, startrow=6)
    psvol.to_excel(writer, sheet_name='Product Dashboard', header=None, index=False,startcol=10, startrow=47)
    psval.to_excel(writer, sheet_name='Product Dashboard', header=None, index=False,startcol=10, startrow=19)
    writer.save()
    
    
    wb = load_workbook("LOS_GC_SUMMARY1.xlsx")
    ws = wb['Summary']
    
    ws['M20'] = '=E20+G20+I20+K20'
    ws['N20'] = '=F20+H20+J20+L20'
    ws['K19'] = 387.7
    ws['L19'] = 352
    ws['K20'] = 387.7
    ws['L20'] = 352
    ws2=wb['Branch Dashboard']
    ws2['T3']=day.strftime("%d-%m-%Y")
    
    ws3=wb['AUM_Product_Trend']
    
    ws3['L16']=387.7
    ws3['L17']='=SUM(L5:L16)'
    wb.save("LOS_GC_SUMMARY1.xlsx")
else:
    print ("DUPLICATION ERROR")
    